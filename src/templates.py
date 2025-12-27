"""Shared template schema and helpers for Data Frame Tool.

The template format captures a consistent set of fields used across the
command-line tools and UIs. All template writers/readers should import from
this module to guarantee compatibility.
"""

from __future__ import annotations  # <--- MUST BE FIRST

import importlib.util
import json
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Check for PyYAML availability
_yaml_available = importlib.util.find_spec("yaml") is not None
if _yaml_available:
    import yaml  # type: ignore
else:
    yaml = None

COMMON_FIELDS_HELP = (
    "Templates record the sheet name, 0-indexed header row, optional skiprows, "
    "selected columns, and optional column renames so any tool can reuse them."
)

LEGACY_NAME_HINT = (
    "No df-template file found. If you have an older '*_template.(json|yml)' "
    "file, pass it explicitly or rename it to '<stem>.df-template.json' to use "
    "the unified format."
)


# --- Helper Functions (Top Level) ---


def parse_skiprows(raw_value: str | None) -> List[int]:
    """Parse a comma-separated string into a list of integers."""
    if not raw_value:
        return []

    values: List[int] = []
    for part in str(raw_value).split(","):
        text = part.strip()
        if not text:
            continue
        try:
            values.append(int(text))
        except ValueError:
            continue
    return values


# --- Data Classes ---


@dataclass
class HeaderCell:
    """Represents the position of a header cell along with its mapping."""

    name: str
    column: int
    row: int
    alias: Optional[str] = None
    is_metadata: bool = False
    metadata_type: str = "header"

    def to_dict(self) -> Dict[str, Any]:
        payload: Dict[str, Any] = {
            "name": self.name,
            "column": self.column,
            "row": self.row,
        }
        if self.alias:
            payload["alias"] = self.alias
        payload["is_metadata"] = bool(self.is_metadata)
        payload["metadata_type"] = self.metadata_type
        return payload

    @classmethod
    def from_dict(cls, payload: dict) -> "HeaderCell":
        return cls(
            name=str(payload["name"]),
            column=int(payload["column"]),
            row=int(payload["row"]),
            alias=payload.get("alias"),
            is_metadata=bool(payload.get("is_metadata", False)),
            metadata_type=str(payload.get("metadata_type", "header")),
        )


@dataclass
class Template:
    """Unified representation of a template."""

    sheet: str | int | None
    sheets: List[str | int] = field(default_factory=list)
    header_row: int = 0
    columns: List[str] = field(default_factory=list)
    column_mappings: Dict[str, str] = field(default_factory=dict)
    headers: List[HeaderCell] = field(default_factory=list)
    skiprows: List[int] = field(default_factory=list)
    delimiter: str = ","
    encoding: str = "utf-8"
    source_type: str = "excel"
    source_file: Optional[str] = None
    output_dir: Optional[str] = None
    provider_name: Optional[str] = None
    combine_sheets: bool = False
    combine_on: List[str] = field(default_factory=list)
    connection_name: Optional[str] = None
    sql_table: Optional[str] = None
    sql_query: Optional[str] = None
    trim_strings: bool = True
    drop_empty_rows: bool = False
    drop_null_columns_threshold: Optional[float] = None
    dedupe_on: List[str] = field(default_factory=list)
    strip_thousands: bool = False
    unpivot: bool = False
    id_columns: List[str] = field(default_factory=list)
    var_name: str = "report_date"
    value_name: str = "sales_amount"
    required_fields: List[str] = field(default_factory=list)
    field_types: Dict[str, str] = field(default_factory=dict)
    template_version: int = 3

    def to_dict(self) -> Dict[str, Any]:
        return {
            "template_version": self.template_version,
            "source_type": self.source_type,
            "sheet": self.sheet,
            "sheets": self.sheets,
            "header_row": self.header_row,
            "skiprows": self.skiprows,
            "delimiter": self.delimiter,
            "encoding": self.encoding,
            "columns": self.columns,
            "column_mappings": self.column_mappings,
            "headers": [h.to_dict() for h in self.headers],
            "source_file": self.source_file,
            "output_dir": self.output_dir,
            "provider_name": self.provider_name,
            "combine_sheets": self.combine_sheets,
            "combine_on": self.combine_on,
            "connection_name": self.connection_name,
            "trim_strings": self.trim_strings,
            "drop_empty_rows": self.drop_empty_rows,
            "drop_null_columns_threshold": self.drop_null_columns_threshold,
            "dedupe_on": self.dedupe_on,
            "strip_thousands": self.strip_thousands,
            "sql_table": self.sql_table,
            "sql_query": self.sql_query,
            "unpivot": self.unpivot,
            "id_columns": self.id_columns,
            "var_name": self.var_name,
            "value_name": self.value_name,
            "required_fields": self.required_fields,
            "field_types": self.field_types,
        }

    @classmethod
    def from_dict(cls, payload: dict) -> "Template":
        if not isinstance(payload, dict):
            raise ValueError("Template file must contain a JSON/YAML object")

        sheet = payload.get("sheet", payload.get("sheet_name"))
        sheets_raw = payload.get("sheets", [])
        sheets: List[str | int] = []
        if isinstance(sheets_raw, list):
            sheets = [s for s in sheets_raw if s is not None]

        header_raw = payload.get("header_row", payload.get("header", 0))
        header_row = int(header_raw) if header_raw is not None else 0

        columns = [
            str(c)
            for c in payload.get("columns", payload.get("selected_headers", []))
            if c is not None
        ]

        mapping_payload = (
            payload.get("column_mappings", payload.get("header_mapping", {})) or {}
        )
        column_mappings = {str(k): str(v) for k, v in mapping_payload.items()}

        headers = [HeaderCell.from_dict(item) for item in payload.get("headers", [])]

        skiprows_raw = payload.get("skiprows", [])
        skiprows = list(skiprows_raw) if isinstance(skiprows_raw, list) else []

        delimiter = payload.get("delimiter", ",")
        encoding = payload.get("encoding", "utf-8")
        source_type = payload.get("source_type", "excel")
        source_file = payload.get("source_file", payload.get("excel_file"))
        output_dir = payload.get("output_dir")
        provider_name = payload.get("provider_name")
        combine_sheets = bool(payload.get("combine_sheets", False))
        combine_on_raw = payload.get("combine_on", [])
        combine_on: List[str] = []
        if isinstance(combine_on_raw, list):
            combine_on = [str(item) for item in combine_on_raw if item not in (None, "")]
        elif isinstance(combine_on_raw, str):
            combine_on = [part.strip() for part in combine_on_raw.split(",") if part.strip()]
        unpivot = bool(payload.get("unpivot", False))

        id_columns_raw = payload.get("id_columns", [])
        id_columns = list(id_columns_raw) if isinstance(id_columns_raw, list) else []

        var_name = payload.get("var_name", "report_date")
        value_name = payload.get("value_name", "sales_amount")
        required_fields_raw = payload.get("required_fields", [])
        required_fields: List[str] = (
            [str(f) for f in required_fields_raw if f not in (None, "")]
            if isinstance(required_fields_raw, list)
            else []
        )
        field_types_raw = payload.get("field_types", {})
        field_types: Dict[str, str] = (
            {str(k): str(v) for k, v in field_types_raw.items()}
            if isinstance(field_types_raw, dict)
            else {}
        )

        template_version = int(payload.get("template_version", 3))
        connection_name = payload.get("connection_name")
        trim_strings = bool(payload.get("trim_strings", True))
        drop_empty_rows = bool(payload.get("drop_empty_rows", False))
        drop_null_columns_threshold_raw = payload.get("drop_null_columns_threshold")
        try:
            drop_null_columns_threshold = (
                float(drop_null_columns_threshold_raw)
                if drop_null_columns_threshold_raw is not None
                else None
            )
        except (TypeError, ValueError):
            drop_null_columns_threshold = None
        dedupe_raw = payload.get("dedupe_on", [])
        dedupe_on: List[str] = []
        if isinstance(dedupe_raw, list):
            dedupe_on = [str(item) for item in dedupe_raw if item not in (None, "")]
        elif isinstance(dedupe_raw, str):
            dedupe_on = [part.strip() for part in dedupe_raw.split(",") if part.strip()]
        strip_thousands = bool(payload.get("strip_thousands", False))
        sql_table = payload.get("sql_table")
        sql_query = payload.get("sql_query")

        # Backwards compatibility for single-sheet templates
        if not sheets and sheet is not None:
            sheets = [sheet]
        if not combine_sheets and len(sheets) > 1:
            combine_sheets = True

        # Backwards compatibility logic
        if headers and not columns:
            columns = [header.name for header in headers]
        if not columns and column_mappings:
            columns = list(column_mappings.keys())

        return cls(
            sheet=sheet,
            sheets=sheets,
            header_row=header_row,
            columns=columns,
            column_mappings=column_mappings,
            headers=headers,
            skiprows=skiprows,
            delimiter=delimiter,
            encoding=encoding,
            source_type=source_type,
            source_file=source_file,
            output_dir=output_dir,
            provider_name=provider_name,
            combine_sheets=combine_sheets,
            combine_on=combine_on,
            connection_name=connection_name,
            trim_strings=trim_strings,
            drop_empty_rows=drop_empty_rows,
            drop_null_columns_threshold=drop_null_columns_threshold,
            dedupe_on=dedupe_on,
            strip_thousands=strip_thousands,
            sql_table=sql_table,
            sql_query=sql_query,
            unpivot=unpivot,
            id_columns=id_columns,
            var_name=var_name,
            value_name=value_name,
            required_fields=required_fields,
            field_types=field_types,
            template_version=template_version,
        )


# --- File I/O Helpers ---


def default_template_path(source: Path, suffix: str = "json") -> Path:
    """Return the standardized template path for a given data file."""
    safe_suffix = suffix.lstrip(".")
    return source.with_name(f"{source.stem}.df-template.{safe_suffix}")


def locate_template(directory: Path, stem: str | None = None) -> Path:
    """Find a template within ``directory`` using the unified naming scheme."""
    candidates: list[Path] = []
    if stem:
        for ext in ("json", "yaml", "yml"):
            path = directory / f"{stem}.df-template.{ext}"
            if path.exists():
                return path
            candidates.append(path)
    else:
        for ext in ("json", "yaml", "yml"):
            matches = sorted(directory.glob(f"*.df-template.{ext}"))
            if matches:
                return matches[0]
            candidates.append(directory / f"<name>.df-template.{ext}")

    legacy = sorted(directory.glob("*_template.*"))
    if legacy:
        raise FileNotFoundError(
            f"{LEGACY_NAME_HINT} Found legacy file: {legacy[0].name}"
        )

    raise FileNotFoundError(
        f"No template found. Expected one of: {', '.join(str(p) for p in candidates)}"
    )


def locate_streamlit_template(directory: Path, stem: str | None = None) -> Path:
    """Find a Streamlit-generated template using the .df-template.json naming scheme."""
    candidates: list[Path] = []
    if stem:
        path = directory / f"{stem}.df-template.json"
        if path.exists():
            return path
        candidates.append(path)
    else:
        matches = sorted(directory.glob("*.df-template.json"))
        if matches:
            return matches[0]
        candidates.append(directory / "<name>.df-template.json")

    raise FileNotFoundError(
        f"No Streamlit template found. Expected one of: {', '.join(str(p) for p in candidates)}"
    )


def _load_payload(path: Path) -> dict:
    if path.suffix.lower() in {".yaml", ".yml"}:
        if not _yaml_available:
            raise ValueError("PyYAML is required to load YAML templates")
        with path.open("r", encoding="utf-8") as handle:
            return yaml.safe_load(handle)

    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def load_template(path: Path) -> Template:
    """Load a template file (JSON or YAML) into a ``Template`` instance."""
    if not path.exists():
        raise FileNotFoundError(f"Template not found: {path}")

    payload = _load_payload(path)
    return Template.from_dict(payload)


def save_template(template: Template, path: Path) -> None:
    """Persist a template to disk using the path's extension."""
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.suffix.lower() in {".yaml", ".yml"}:
        if not _yaml_available:
            raise ValueError("PyYAML is required to save YAML templates")
        with path.open("w", encoding="utf-8") as handle:
            yaml.safe_dump(template.to_dict(), handle, sort_keys=False)
    else:
        with path.open("w", encoding="utf-8") as handle:
            json.dump(template.to_dict(), handle, indent=2)


def describe_common_fields() -> str:
    """Return human-readable text describing the unified template format."""
    return COMMON_FIELDS_HELP


# --- Header Normalization Helpers ---


def _effective_header_row(
    header_row: int, skiprows: Optional[List[int]] | None = None
) -> int:
    """Translate a pandas-style header row into a 1-indexed worksheet row."""
    skipped = skiprows or []
    skipped_before = len(
        [row for row in skipped if isinstance(row, int) and row <= header_row]
    )
    return header_row + skipped_before + 1


def normalize_excel_headers(
    path: Path,
    sheet: str | int | None,
    header_row: int,
    skiprows: Optional[List[int]] | None = None,
) -> Tuple[List[str], bool]:
    """Return normalized headers for a worksheet, expanding merged regions."""

    workbook = load_workbook(path, read_only=False, data_only=False)
    if isinstance(sheet, int):
        try:
            worksheet = workbook.worksheets[sheet]
        except IndexError:
            # Fallback if sheet index is out of range
            worksheet = workbook.active
    elif sheet is None:
        worksheet = workbook.active
    else:
        try:
            worksheet = workbook[sheet]
        except KeyError:
            # Fallback
            worksheet = workbook.active

    target_row = _effective_header_row(header_row, skiprows)
    row_cells = list(
        worksheet.iter_rows(min_row=target_row, max_row=target_row, values_only=False)
    )

    if not row_cells:
        return [], False

    headers = ["" if cell.value is None else str(cell.value) for cell in row_cells[0]]

    merged_detected = False
    for merged_range in worksheet.merged_cells.ranges:
        if not (merged_range.min_row <= target_row <= merged_range.max_row):
            continue

        merged_detected = True
        base_value = worksheet.cell(merged_range.min_row, merged_range.min_col).value

        placeholder_base = (
            str(base_value)
            if base_value not in (None, "")
            else f"merged_{get_column_letter(merged_range.min_col)}{merged_range.min_row}"
        )

        for col in range(merged_range.min_col, merged_range.max_col + 1):
            header_value = placeholder_base
            # If spanning multiple columns, append _B, _C etc to make unique
            if base_value in (None, "") and merged_range.max_col > merged_range.min_col:
                header_value = f"{placeholder_base}_{get_column_letter(col)}"

            index = col - 1
            while len(headers) <= index:
                headers.append("")
            headers[index] = header_value

    return headers, merged_detected


def apply_normalized_headers(df: pd.DataFrame, headers: List[str]) -> pd.DataFrame:
    """Return a copy of ``df`` whose columns are replaced with ``headers``."""
    if not headers:
        return df

    normalized = list(headers)
    if len(normalized) < len(df.columns):
        normalized.extend(str(col) for col in df.columns[len(normalized) :])
    elif len(normalized) > len(df.columns):
        normalized = normalized[: len(df.columns)]

    adjusted = df.copy()
    adjusted.columns = normalized
    return adjusted


def filter_and_rename(df: pd.DataFrame, template: Template) -> pd.DataFrame:
    """Subset and rename columns using a template."""
    if template.headers:
        rename_map: dict[str, str] = {}
        for idx, header in enumerate(template.headers):
            if idx >= len(df.columns):
                continue
            current_name = df.columns[idx]
            target = header.alias or template.column_mappings.get(
                header.name, header.name
            )
            rename_map[current_name] = target

        # Slice only mapped columns
        df = df.iloc[:, : len(rename_map)] if rename_map else df
        if rename_map:
            df = df.rename(columns=rename_map)
    else:
        if template.columns:
            available = [col for col in template.columns if col in df.columns]
            df = df[available]

        if template.column_mappings:
            rename_map = {
                col: template.column_mappings.get(col, col) for col in df.columns
            }
            df = df.rename(columns=rename_map)

    return df


def read_excel_with_template(path: Path, template: Template) -> pd.DataFrame:
    """Read an Excel or CSV file using template settings, optionally combining sheets."""

    logger = logging.getLogger(__name__)

    # CSV Handling
    if path.suffix.lower() == ".csv" or template.source_type == "csv":
        df = pd.read_csv(
            path,
            header=template.header_row,
            skiprows=template.skiprows,
            sep=template.delimiter,
            encoding=template.encoding,
        )
        return filter_and_rename(df, template)

    # Excel Handling (with optional multi-sheet combine)
    if template.combine_sheets and template.sheets:
        sheet_list: list[str | int] = list(template.sheets)
    elif template.sheet is not None:
        sheet_list = [template.sheet]
    else:
        sheet_list = [0]

    frames: list[pd.DataFrame] = []
    any_merged = False

    for sheet in sheet_list:
        normalized_headers: list[str] = []
        merged_detected = False
        try:
            normalized_headers, merged_detected = normalize_excel_headers(
                path=path,
                sheet=sheet,
                header_row=template.header_row,
                skiprows=template.skiprows,
            )
        except Exception as exc:
            logger.debug("Header normalization skipped for sheet %s: %s", sheet, exc)

        if template.headers:
            use_columns = [cell.column for cell in template.headers]
        else:
            use_columns = template.columns or None

        df = pd.read_excel(
            path,
            sheet_name=sheet if sheet is not None else 0,
            header=template.header_row,
            skiprows=template.skiprows,
            usecols=use_columns,
        )

        if normalized_headers:
            df = apply_normalized_headers(df, normalized_headers)

        df = df.dropna(how="all").dropna(axis=1, how="all")

        df = filter_and_rename(df, template)

        if template.combine_sheets:
            df["source_sheet"] = str(sheet)
        frames.append(df)
        any_merged = any_merged or merged_detected

    if not frames:
        return pd.DataFrame()

    result = pd.concat(frames, ignore_index=True) if len(frames) > 1 else frames[0]

    if any_merged:
        logger.warning("Merged header cells detected in %s.", path.name)

    return result
