"""Tkinter application to build templates from tabular files."""

from __future__ import annotations

import os
import json
import tkinter as tk
from pathlib import Path
from tkinter import filedialog, messagebox, simpledialog, ttk
from typing import Dict, List, Optional
import threading

import pandas as pd
from openpyxl import load_workbook

from .core import DEFAULT_SCHEMA_CANDIDATES, SCHEMA_DIR, TARGET_SCHEMA
from .services.header_detection import get_normalized_headers, guess_header_row
from .services.io import read_preview_frame, sheet_names
from .services.mapping import (
    auto_map_columns,
    describe_schema,
    learn_synonyms_from_mapping,
    load_target_schema,
    snake_case,
)
from .services.schema_candidates import build_schema_candidates
from .templates import (
    HeaderCell,
    Template,
    default_template_path,
    describe_common_fields,
    load_template,
    apply_normalized_headers,
    parse_skiprows,
    save_template,
    _yaml_available,
)
from .connectors import (
    ConnectionConfig,
    fetch_sql_preview,
    load_connections,
    save_connections,
    test_connection,
)
from .combine_runner import run_combine, read_frame

DEFAULT_PREVIEW_ROWS = 10


class ExcelTemplateApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("Excel Ingestor Pro")
        self.root.geometry("1200x900")

        # Data State
        self.file_path: Optional[str] = None
        self.sheet_names: List[str] = []
        self.columns: List[str] = []
        self.mapping: Dict[str, str] = {}
        self.header_cells: Dict[str, HeaderCell] = {}
        self.preview_df: Optional[pd.DataFrame] = None
        self.selection_mode = tk.StringVar(value="preview")
        self.selected_metadata_cells: set[tuple[int, int]] = set()
        self.metadata_cells: List[HeaderCell] = []
        self.metadata_listbox: Optional[tk.Listbox] = None

        # Load Schema for Dropdowns
        self.target_schema = load_target_schema()
        self.target_fields = list(self.target_schema.keys())
        self.current_schema_path = self._guess_schema_path()
        self.schema_path_var = tk.StringVar(value=str(self.current_schema_path))
        self.schema_fields_var = tk.StringVar(
            value=", ".join(self.target_fields) if self.target_fields else "-"
        )
        self.schema_source_path: Optional[str] = None
        self.schema_header_row_var = tk.StringVar(value="0")
        self.schema_skiprows_var = tk.StringVar(value="")
        self.schema_delimiter_var = tk.StringVar(value=",")
        self.schema_encoding_var = tk.StringVar(value="utf-8")
        self.schema_data_type_var = tk.StringVar(value="generic")
        self.schema_candidates: List[Dict[str, object]] = []
        self.suggestion_detail_var = tk.StringVar(value="")
        self.custom_map_path = SCHEMA_DIR / "header_map.txt"
        self.custom_map_label_var = tk.StringVar(value=str(self.custom_map_path))
        self.final_diff_missing_var = tk.StringVar(value="")
        self.final_diff_extra_var = tk.StringVar(value="")
        self.final_missing_set: set[str] = set()
        self.final_extra_set: set[str] = set()
        self.tooltip_label: Optional[tk.Label] = None
        self.status_var = tk.StringVar(value="Ready")
        self._busy = False

        # UI Variables
        self.root.bind("<Control-s>", lambda _e: self.save_template())
        self.sheet_var = tk.StringVar()
        self.header_row_var = tk.StringVar(value="0")
        self.skiprows_var = tk.StringVar(value="")
        self.delimiter_var = tk.StringVar(value=",")
        self.encoding_var = tk.StringVar(value="utf-8")
        self.combine_sheets_var = tk.BooleanVar(value=False)
        self.connection_name_var = tk.StringVar(value="")

        # Transformation Variables
        self.alias_var = tk.StringVar()
        self.format_var = tk.StringVar(value="json")
        self.unpivot_var = tk.BooleanVar(value=False)
        self.var_name_var = tk.StringVar(value="report_date")
        self.val_name_var = tk.StringVar(value="amount")
        self.combine_on_var = tk.StringVar(value="")
        self.trim_strings_var = tk.BooleanVar(value=True)
        self.drop_empty_rows_var = tk.BooleanVar(value=False)
        self.drop_null_threshold_var = tk.StringVar(value="")
        self.dedupe_on_var = tk.StringVar(value="")
        self.strip_thousands_var = tk.BooleanVar(value=False)
        self.sql_table_var = tk.StringVar(value="")
        self.sql_query_var = tk.StringVar(value="")
        self.combine_mode_var = tk.StringVar(value="concat")
        self.combine_keys_var = tk.StringVar(value="")
        self.combine_how_var = tk.StringVar(value="inner")
        self.combine_strict_var = tk.BooleanVar(value=False)
        self.combine_pattern_var = tk.StringVar(value="*.xlsx")
        self.combine_output_var = tk.StringVar(value="data/output/Master_Sales_Report.xlsx")
        self.combine_input_dir_var = tk.StringVar(value="data/output")
        # Mapping state snapshot for reset
        self.saved_schema_snapshot: Dict[str, List[str]] | None = None

        # Connection store (in-memory)
        self.connections: List[ConnectionConfig] = load_connections()

        self.source_type: str = "excel"
        self._build_ui()
        self._refresh_connection_list()
        self._init_styles()

    def _build_ui(self) -> None:
        notebook = ttk.Notebook(self.root)
        import_tab = ttk.Frame(notebook)
        schema_tab = ttk.Frame(notebook)
        process_tab = ttk.Frame(notebook)
        save_tab = ttk.Frame(notebook)
        # Order tabs so schema selection comes first
        notebook.add(schema_tab, text="Schema")
        notebook.add(import_tab, text="Import Data")
        notebook.add(process_tab, text="Process & Map")
        notebook.add(save_tab, text="Save & Validate")
        notebook.pack(fill="both", expand=True)

        # --- Schema Pane with scrollable passes ---
        schema_canvas = tk.Canvas(schema_tab)
        schema_scroll = ttk.Scrollbar(schema_tab, orient="vertical", command=schema_canvas.yview)
        schema_frame = ttk.Frame(schema_canvas)
        schema_frame.bind(
            "<Configure>", lambda _e: schema_canvas.configure(scrollregion=schema_canvas.bbox("all"))
        )
        canvas_window = schema_canvas.create_window((0, 0), window=schema_frame, anchor="nw")
        schema_canvas.configure(yscrollcommand=schema_scroll.set)
        schema_canvas.pack(side="left", fill="both", expand=True)
        schema_scroll.pack(side="right", fill="y")

        # Keep canvas width synced
        def _resize_canvas(event):
            schema_canvas.itemconfig(canvas_window, width=event.width)

        schema_canvas.bind("<Configure>", _resize_canvas)

        # Pass 0: overview and file/schema controls
        schema_info = ttk.LabelFrame(schema_frame, text="Target Schema")
        schema_info.pack(fill="x", padx=10, pady=8)
        path_row = ttk.Frame(schema_info)
        path_row.pack(fill="x", padx=6, pady=4)
        ttk.Label(path_row, text="Current schema file:").pack(side="left")
        ttk.Label(
            path_row,
            textvariable=self.schema_path_var,
            foreground="gray",
        ).pack(side="left", padx=6)
        ttk.Button(
            path_row, text="Choose file...", command=self._choose_schema_file
        ).pack(side="right", padx=4)
        ttk.Button(
            path_row, text="Open folder", command=self._open_schema_folder
        ).pack(side="right", padx=4)
        ttk.Button(
            path_row, text="Reload default", command=lambda: self._reload_target_schema()
        ).pack(side="right", padx=4)
        ttk.Button(
            path_row,
            text="Reset to built-in",
            command=self._reset_to_builtin_schema,
        ).pack(side="right", padx=4)

        fields_frame = ttk.LabelFrame(schema_frame, text="Fields & Synonyms")
        fields_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        ttk.Label(
            fields_frame,
            text="These fields populate the mapping dropdowns. Update the schema file to change them.",
            foreground="gray",
        ).pack(anchor="w", padx=8, pady=(6, 2))
        btn_row = ttk.Frame(fields_frame)
        btn_row.pack(fill="x", padx=8, pady=(0, 6))
        ttk.Button(
            btn_row, text="Load schema from Excel/CSV...", command=self._load_schema_from_excel
        ).pack(side="left", padx=(0, 6))
        self._attach_tooltip(btn_row.winfo_children()[-1], "Pick a sample output file to derive headers and generate a schema.")
        ttk.Button(
            btn_row, text="Save schema...", command=self._save_schema_file
        ).pack(side="left", padx=(0, 6))
        self._attach_tooltip(btn_row.winfo_children()[-1], "Persist current schema fields/synonyms to JSON (default data/schemas/schema.json).")
        ttk.Button(
            btn_row, text="+ Custom mapping", command=self._open_custom_map_editor
        ).pack(side="left", padx=(0, 6))
        ttk.Label(
            btn_row,
            textvariable=self.custom_map_label_var,
            foreground="gray",
        ).pack(side="left", padx=(6, 0))
        ttk.Label(
            fields_frame,
            text="Workflow: 1) Pick/load schema file  2) (Optional) Load sample output to derive headers  3) Apply mapping or heuristics  4) Save schema.",
            foreground="gray",
            wraplength=900,
        ).pack(anchor="w", padx=8, pady=(0, 4))
        ttk.Label(
            fields_frame,
            text="Tip: Custom mapping lets you rename headers (e.g., 'prod code -> product_id'). It applies to the current schema; save if you want it persisted.",
            foreground="gray",
            wraplength=900,
        ).pack(anchor="w", padx=8, pady=(2, 6))
        ttk.Label(
            fields_frame,
            text="Hover over data type, heuristics, or mapping buttons for reminders.",
            foreground="gray",
        ).pack(anchor="w", padx=8, pady=(0, 4))
        ttk.Label(
            fields_frame, textvariable=self.schema_fields_var, wraplength=900
        ).pack(anchor="w", padx=8, pady=(0, 8))
        self.schema_listbox = tk.Listbox(fields_frame, height=10)
        self.schema_listbox.pack(fill="both", expand=True, padx=8, pady=(0, 8))
        # Mapping quick actions
        mapping_row = ttk.Frame(fields_frame)
        mapping_row.pack(fill="x", padx=8, pady=(0, 6))
        ttk.Button(mapping_row, text="Apply mapping + Save schema", command=self._apply_and_save_schema).pack(side="left", padx=(0, 6))
        ttk.Button(mapping_row, text="Reset mapping to defaults", command=self._reset_mapping_to_defaults).pack(side="left", padx=(0, 6))
        self._attach_tooltip(mapping_row.winfo_children()[0], "Apply current headers/mapping and save to the active schema file.")
        self._attach_tooltip(mapping_row.winfo_children()[1], "Restore the last saved schema fields.")

        # Pass 1: Source configuration and preview for schema derivation
        schema_src = ttk.LabelFrame(schema_frame, text="Pass 1: Source & Preview")
        schema_src.pack(fill="x", padx=10, pady=6)

        top_row = ttk.Frame(schema_src)
        top_row.pack(fill="x", padx=6, pady=4)
        ttk.Label(top_row, text="Data type:").pack(side="left")
        self.data_type_combo = ttk.Combobox(
            top_row,
            textvariable=self.schema_data_type_var,
            values=["generic", "product_sales", "product_descriptions", "sales"],
            state="readonly",
            width=22,
        )
        self.data_type_combo.pack(side="left", padx=6)
        ttk.Button(top_row, text="Apply heuristics", command=lambda: self._reload_schema_preview(update_schema=False, use_heuristics=True)).pack(side="right", padx=4)
        ttk.Label(
            top_row,
            text="(Choose a hint to bias header suggestions)",
            foreground="gray",
        ).pack(side="left", padx=6)
        self._attach_tooltip(self.data_type_combo, "Bias header suggestions toward typical shapes: product_sales, product_descriptions, or generic.")

        schema_opt_row = ttk.Frame(schema_src)
        schema_opt_row.pack(fill="x", padx=6, pady=4)

        sheets_col = ttk.Frame(schema_opt_row)
        sheets_col.pack(side="left", padx=(0, 12))
        ttk.Label(sheets_col, text="Sheets:").pack(anchor="w")
        self.schema_sheet_listbox = tk.Listbox(
            sheets_col, height=5, selectmode="extended", exportselection=False
        )
        sheet_scroll = ttk.Scrollbar(
            sheets_col, orient="vertical", command=self.schema_sheet_listbox.yview
        )
        self.schema_sheet_listbox.config(yscrollcommand=sheet_scroll.set)
        self.schema_sheet_listbox.pack(side="left")
        sheet_scroll.pack(side="left", fill="y")

        opt_inputs = ttk.Frame(schema_opt_row)
        opt_inputs.pack(side="left", padx=5, pady=2)

        row_one = ttk.Frame(opt_inputs)
        row_one.pack(fill="x")
        ttk.Label(row_one, text="Header Row:").pack(side="left")
        ttk.Entry(row_one, textvariable=self.schema_header_row_var, width=5).pack(
            side="left", padx=(2, 12)
        )

        ttk.Label(row_one, text="Skip Rows:").pack(side="left")
        ttk.Entry(row_one, textvariable=self.schema_skiprows_var, width=10).pack(
            side="left", padx=(2, 12)
        )

        ttk.Label(row_one, text="Delimiter:").pack(side="left")
        ttk.Entry(row_one, textvariable=self.schema_delimiter_var, width=6).pack(
            side="left", padx=(2, 12)
        )

        ttk.Label(row_one, text="Encoding:").pack(side="left")
        ttk.Entry(row_one, textvariable=self.schema_encoding_var, width=12).pack(
            side="left", padx=(2, 12)
        )

        row_two = ttk.Frame(opt_inputs)
        row_two.pack(fill="x", pady=(6, 0))
        ttk.Button(row_two, text="Reload Preview", command=lambda: self._reload_schema_preview(update_schema=True)).pack(
            side="left", padx=(0, 8)
        )
        ttk.Button(row_two, text="Reset View", command=self._reset_schema_preview).pack(
            side="left"
        )

        schema_prev = ttk.LabelFrame(schema_frame, text="Pass 1 Preview")
        schema_prev.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        self.schema_preview_tree = ttk.Treeview(schema_prev, show="headings", height=10)
        self.schema_preview_tree.pack(fill="both", expand=True, padx=5, pady=5)

        # Pass 2: Heuristic suggestions and selection
        suggestions_frame = ttk.LabelFrame(schema_frame, text="Pass 2: Suggested headers")
        suggestions_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        ttk.Label(
            suggestions_frame,
            text="Pick a candidate header set derived from the data and heuristics.",
            foreground="gray",
        ).pack(anchor="w", padx=8, pady=(6, 2))
        ttk.Label(
            suggestions_frame,
            text="Hover/select to see score and missing/extra vs current schema.",
            foreground="gray",
        ).pack(anchor="w", padx=8, pady=(0, 4))
        self.suggestions_listbox = tk.Listbox(suggestions_frame, height=6, exportselection=False)
        self.suggestions_listbox.pack(fill="both", expand=True, padx=8, pady=(0, 6))
        self.suggestions_listbox.bind("<<ListboxSelect>>", lambda _e: self._on_candidate_select())
        detail_frame = ttk.Frame(suggestions_frame)
        detail_frame.pack(fill="x", padx=8, pady=(0, 6))
        ttk.Label(
            detail_frame,
            text="Details:",
            foreground="gray",
        ).pack(side="left")
        ttk.Label(
            detail_frame,
            textvariable=self.suggestion_detail_var,
            wraplength=900,
            foreground="gray",
            anchor="w",
            justify="left",
        ).pack(side="left", fill="x", expand=True, padx=(4, 0))
        ttk.Button(
            suggestions_frame,
            text="Apply selected candidate",
            command=self._apply_candidate_schema,
        ).pack(anchor="e", padx=8, pady=(0, 8))

        # Pass 3: Final preview with cleaned headers
        final_frame = ttk.LabelFrame(schema_frame, text="Pass 3: Final header preview")
        final_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))
        self.final_preview_tree = ttk.Treeview(final_frame, show="headings", height=10)
        self.final_preview_tree.pack(fill="both", expand=True, padx=5, pady=5)
        diff_row = ttk.Frame(final_frame)
        diff_row.pack(fill="x", padx=5, pady=(2, 4))
        ttk.Label(diff_row, text="Missing:", foreground="red").pack(side="left")
        ttk.Label(diff_row, textvariable=self.final_diff_missing_var, foreground="red").pack(side="left", padx=(4, 12))
        ttk.Label(diff_row, text="Extra:", foreground="blue").pack(side="left")
        ttk.Label(diff_row, textvariable=self.final_diff_extra_var, foreground="blue").pack(side="left", padx=(4, 0))

        # --- Connections ---
        conn_frame = ttk.LabelFrame(import_tab, text="Connections")
        conn_frame.pack(fill="x", padx=10, pady=5)
        ttk.Label(
            conn_frame,
            text="Select an existing connection or use local file import.",
            foreground="gray",
        ).pack(anchor="w", padx=5, pady=(2, 4))
        conn_row = ttk.Frame(conn_frame)
        conn_row.pack(fill="x", padx=5, pady=2)
        self.connection_listbox = tk.Listbox(
            conn_row, height=3, exportselection=False, width=40
        )
        self.connection_listbox.pack(side="left", fill="x", expand=True)
        self.connection_listbox.bind("<<ListboxSelect>>", lambda _e: self._on_connection_select())
        ttk.Button(
            conn_row, text="Add / Edit Connection", command=self.open_connection_manager
        ).pack(side="left", padx=6)
        ttk.Button(
            conn_row, text="Use Connection", command=self.preview_connection
        ).pack(side="left")
        ttk.Button(
            conn_row, text="Test Connection", command=self.test_selected_connection
        ).pack(side="left", padx=4)
        ttk.Button(
            conn_row, text="Use Local File", command=self.select_file
        ).pack(side="left", padx=4)

        # --- 1. File Selection ---
        file_frame = ttk.LabelFrame(import_tab, text="Source Configuration")
        file_frame.pack(fill="x", padx=10, pady=5)

        top_row = ttk.Frame(file_frame)
        top_row.pack(fill="x", padx=5, pady=5)
        self.file_label = ttk.Label(
            top_row, text="No file selected", font=("Segoe UI", 9, "italic")
        )
        self.file_label.pack(side="left", padx=10)
        ttk.Label(top_row, text="(or use connection above)").pack(side="left", padx=6)

        # Parsing Options Row
        opt_row = ttk.Frame(file_frame)
        opt_row.pack(fill="x", padx=5, pady=5)

        sheets_col = ttk.Frame(opt_row)
        sheets_col.pack(side="left", padx=(0, 12))
        ttk.Label(sheets_col, text="Sheets:").pack(anchor="w")
        self.sheet_listbox = tk.Listbox(
            sheets_col,
            height=6,
            selectmode="extended",
            exportselection=False,
        )
        sheet_scroll = ttk.Scrollbar(
            sheets_col, orient="vertical", command=self.sheet_listbox.yview
        )
        self.sheet_listbox.config(yscrollcommand=sheet_scroll.set)
        self.sheet_listbox.pack(side="left")
        sheet_scroll.pack(side="left", fill="y")
        ttk.Checkbutton(
            sheets_col,
            text="Combine selected sheets",
            variable=self.combine_sheets_var,
            command=self.load_headers,
        ).pack(anchor="w", pady=(4, 0))

        opt_inputs = ttk.Frame(opt_row)
        opt_inputs.pack(side="left", padx=5, pady=2)

        row_one = ttk.Frame(opt_inputs)
        row_one.pack(fill="x")
        ttk.Label(row_one, text="Header Row:").pack(side="left")
        ttk.Entry(row_one, textvariable=self.header_row_var, width=5).pack(
            side="left", padx=(2, 12)
        )
        ttk.Button(row_one, text="-", command=self._decrement_header, width=2).pack(
            side="left", padx=(0, 4)
        )
        ttk.Button(row_one, text="+", command=self._increment_header, width=2).pack(
            side="left", padx=(0, 12)
        )

        ttk.Label(row_one, text="Skip Rows:").pack(side="left")
        ttk.Entry(row_one, textvariable=self.skiprows_var, width=10).pack(
            side="left", padx=(2, 12)
        )

        ttk.Label(row_one, text="Delimiter:").pack(side="left")
        ttk.Entry(row_one, textvariable=self.delimiter_var, width=6).pack(
            side="left", padx=(2, 12)
        )

        ttk.Label(row_one, text="Encoding:").pack(side="left")
        ttk.Entry(row_one, textvariable=self.encoding_var, width=12).pack(
            side="left", padx=(2, 12)
        )

        row_two = ttk.Frame(opt_inputs)
        row_two.pack(fill="x", pady=(6, 0))
        ttk.Button(row_two, text="Reload Preview", command=self.load_headers).pack(
            side="left", padx=(0, 8)
        )
        ttk.Button(row_two, text="Reset View", command=self.reset_view).pack(
            side="left"
        )

        mode_frame = ttk.LabelFrame(file_frame, text="Selection Mode")
        mode_frame.pack(fill="x", padx=5, pady=(0, 6))
        ttk.Label(mode_frame, text="Action:").pack(side="left", padx=5)
        ttk.Radiobutton(
            mode_frame,
            text="Preview/Process",
            variable=self.selection_mode,
            value="preview",
            command=self._set_selection_mode,
        ).pack(side="left")
        ttk.Radiobutton(
            mode_frame,
            text="Define Metadata",
            variable=self.selection_mode,
            value="metadata",
            command=self._set_selection_mode,
        ).pack(side="left", padx=10)

        info_row = ttk.Frame(file_frame)
        info_row.pack(fill="x", padx=5, pady=(2, 6))
        self.info_vars = {
            "sheets": tk.StringVar(value="Sheets: -"),
            "rows": tk.StringVar(value="Rows: -"),
            "cols": tk.StringVar(value="Columns: -"),
        }
        ttk.Label(info_row, textvariable=self.info_vars["sheets"]).pack(side="left")
        ttk.Label(info_row, textvariable=self.info_vars["rows"], padding=(12, 0)).pack(
            side="left"
        )
        ttk.Label(info_row, textvariable=self.info_vars["cols"], padding=(12, 0)).pack(
            side="left"
        )
        status_row = ttk.Frame(file_frame)
        status_row.pack(fill="x", padx=5, pady=(0, 4))
        ttk.Label(status_row, textvariable=self.status_var, foreground="gray").pack(
            side="left"
        )
        # SQL specific controls
        sql_row = ttk.Frame(import_tab)
        sql_row.pack(fill="x", padx=10, pady=(0, 6))
        ttk.Label(sql_row, text="SQL Table:").pack(side="left")
        ttk.Entry(sql_row, textvariable=self.sql_table_var, width=24).pack(
            side="left", padx=4
        )
        ttk.Label(sql_row, text="or SQL Query:").pack(side="left", padx=(10, 2))
        ttk.Entry(sql_row, textvariable=self.sql_query_var, width=50).pack(
            side="left", padx=4
        )
        ttk.Label(sql_row, text="Preview rows: use 'Use Connection'").pack(
            side="left", padx=8
        )

        # --- 2. Data Preview ---
        prev_frame = ttk.LabelFrame(import_tab, text="Data Preview")
        prev_frame.pack(fill="both", expand=False, padx=10, pady=5)

        self.preview_tree = ttk.Treeview(prev_frame, show="headings", height=14)
        self.preview_tree.pack(side="left", fill="both", expand=False, padx=5, pady=5)
        self.preview_tree.bind("<<TreeviewSelect>>", self.on_preview_row_select)
        self.preview_tree.bind("<ButtonRelease-1>", self._on_cell_click)

        meta_container = ttk.Frame(prev_frame)
        meta_container.pack(side="left", fill="y", padx=(10, 5), pady=5)

        ttk.Label(meta_container, text="Selected Metadata:").pack(anchor="w")

        meta_scroll = ttk.Scrollbar(meta_container, orient="vertical")
        self.metadata_listbox = tk.Listbox(
            meta_container,
            height=10,
            width=40,
            exportselection=False,
            yscrollcommand=meta_scroll.set,
        )
        self.metadata_listbox.pack(side="left", fill="both", expand=True)
        meta_scroll.config(command=self.metadata_listbox.yview)
        meta_scroll.pack(side="right", fill="y")

        # --- 3. Mapping & Transformation ---
        map_frame = ttk.LabelFrame(process_tab, text="Mapping & Transformation")
        map_frame.pack(fill="both", expand=True, padx=10, pady=5)

        # Left: Source Columns
        col_frame = ttk.Frame(map_frame)
        col_frame.pack(side="left", fill="y", padx=5, pady=5)
        ttk.Label(col_frame, text="Available Columns:").pack(anchor="w")
        self.columns_listbox = tk.Listbox(
            col_frame, width=30, height=12, selectmode="extended"
        )
        self.columns_listbox.pack(fill="both", expand=True)

        # Center: Mapping Controls
        center_frame = ttk.Frame(map_frame)
        center_frame.pack(side="left", fill="y", padx=10, pady=20)

        ttk.Label(center_frame, text="Map to Target Field:").pack()
        # STRICT DROPDOWN ENFORCEMENT
        self.alias_combo = ttk.Combobox(
            center_frame,
            textvariable=self.alias_var,
            values=self.target_fields,
            state="readonly",
        )
        self.alias_combo.pack(pady=5)

        ttk.Button(center_frame, text="Add Mapping >>", command=self.add_mapping).pack(
            pady=5
        )
        ttk.Button(center_frame, text="<< Remove", command=self.remove_mapping).pack(
            pady=5
        )
        ttk.Separator(center_frame, orient="horizontal").pack(fill="x", pady=10)
        ttk.Button(
            center_frame, text="Auto-Suggest", command=self.apply_smart_mapping
        ).pack(pady=5)

        # Right: Mapped Configuration
        right_frame = ttk.Frame(map_frame)
        right_frame.pack(side="right", fill="both", expand=True, padx=5, pady=5)
        header_row = ttk.Frame(right_frame)
        header_row.pack(fill="x")
        ttk.Label(header_row, text="Configuration:").pack(anchor="w", side="left")
        ttk.Button(
            header_row, text="Save Template", command=self.save_template
        ).pack(side="right", padx=4)
        ttk.Button(header_row, text="Validate", command=self.run_validation).pack(
            side="right", padx=4
        )

        self.mapping_tree = ttk.Treeview(
            right_frame, columns=("src", "target", "role"), show="headings", height=8
        )
        self.mapping_tree.heading("src", text="Source")
        self.mapping_tree.heading("target", text="Target Field")
        self.mapping_tree.heading("role", text="Role")  # Identifier or Data
        self.mapping_tree.column("src", width=120)
        self.mapping_tree.column("target", width=120)
        self.mapping_tree.column("role", width=80)
        self.mapping_tree.pack(fill="both", expand=True)

        # --- 4. Unpivot / Melt Options ---
        trans_frame = ttk.LabelFrame(process_tab, text="Structure Transformation")
        trans_frame.pack(fill="x", padx=10, pady=5)

        chk_unpivot = ttk.Checkbutton(
            trans_frame,
            text="Unpivot (Melt) Data?",
            variable=self.unpivot_var,
            command=self._refresh_mapping_view,
        )
        chk_unpivot.pack(side="left", padx=10)

        ttk.Label(trans_frame, text="Variable Name (e.g. Month):").pack(
            side="left", padx=(10, 2)
        )
        ttk.Entry(trans_frame, textvariable=self.var_name_var, width=15).pack(
            side="left"
        )

        ttk.Label(trans_frame, text="Value Name (e.g. Amount):").pack(
            side="left", padx=(10, 2)
        )
        ttk.Entry(trans_frame, textvariable=self.val_name_var, width=15).pack(
            side="left"
        )

        ttk.Label(trans_frame, text="Group by (comma-separated):").pack(
            side="left", padx=(10, 2)
        )
        self.combine_on_combo = ttk.Combobox(
            trans_frame,
            textvariable=self.combine_on_var,
            values=[""] + self.target_fields,
            state="normal",  # allow typing multiple fields
            width=24,
        )
        self.combine_on_combo.pack(side="left")

        # --- Cleanup Options ---
        clean_frame = ttk.LabelFrame(process_tab, text="Cleanup Options")
        clean_frame.pack(fill="x", padx=10, pady=5)

        ttk.Checkbutton(
            clean_frame, text="Trim text fields", variable=self.trim_strings_var
        ).pack(side="left", padx=8, pady=4)
        ttk.Checkbutton(
            clean_frame,
            text="Drop fully empty rows",
            variable=self.drop_empty_rows_var,
        ).pack(side="left", padx=8, pady=4)
        ttk.Checkbutton(
            clean_frame,
            text="Strip thousands separators in text columns",
            variable=self.strip_thousands_var,
        ).pack(side="left", padx=8, pady=4)

        thresh_frame = ttk.Frame(clean_frame)
        thresh_frame.pack(side="left", padx=8)
        ttk.Label(thresh_frame, text="Drop columns if non-null ratio <").pack(
            side="left"
        )
        ttk.Entry(thresh_frame, textvariable=self.drop_null_threshold_var, width=6).pack(
            side="left", padx=4
        )

        dedupe_frame = ttk.Frame(clean_frame)
        dedupe_frame.pack(side="left", padx=8)
        ttk.Label(dedupe_frame, text="Dedupe on keys:").pack(side="left")
        ttk.Entry(dedupe_frame, textvariable=self.dedupe_on_var, width=24).pack(
            side="left", padx=4
        )

        # --- Combine Options ---
        combine_frame = ttk.LabelFrame(process_tab, text="Combine Options")
        combine_frame.pack(fill="x", padx=10, pady=5)
        ttk.Label(combine_frame, text="Mode:").pack(side="left", padx=4)
        ttk.Combobox(
            combine_frame,
            textvariable=self.combine_mode_var,
            values=["concat", "merge"],
            state="readonly",
            width=10,
        ).pack(side="left", padx=4)
        ttk.Label(combine_frame, text="Keys (comma):").pack(side="left", padx=4)
        ttk.Entry(combine_frame, textvariable=self.combine_keys_var, width=24).pack(
            side="left", padx=4
        )
        ttk.Button(
            combine_frame,
            text="Use mapped fields",
            command=self._use_mapped_keys,
        ).pack(side="left", padx=4)
        ttk.Label(combine_frame, text="Join type:").pack(side="left", padx=4)
        ttk.Combobox(
            combine_frame,
            textvariable=self.combine_how_var,
            values=["inner", "outer", "left", "right"],
            state="readonly",
            width=8,
        ).pack(side="left", padx=4)
        ttk.Checkbutton(
            combine_frame,
            text="Strict schema (concat)",
            variable=self.combine_strict_var,
        ).pack(side="left", padx=8)
        ttk.Label(combine_frame, text="Input dir:").pack(side="left", padx=4)
        ttk.Entry(combine_frame, textvariable=self.combine_input_dir_var, width=20).pack(
            side="left", padx=4
        )
        ttk.Label(combine_frame, text="Pattern:").pack(side="left", padx=4)
        self.combine_pattern_var = tk.StringVar(value="*.xlsx")
        ttk.Entry(combine_frame, textvariable=self.combine_pattern_var, width=10).pack(
            side="left", padx=4
        )
        ttk.Label(combine_frame, text="Output file:").pack(side="left", padx=4)
        self.combine_output_var = tk.StringVar(value="Master_Sales_Report.xlsx")
        ttk.Entry(combine_frame, textvariable=self.combine_output_var, width=24).pack(
            side="left", padx=4
        )
        ttk.Label(
            combine_frame,
            text="Keys must be canonical mapped names (e.g., order_id, article_sku).",
            foreground="gray",
        ).pack(side="left", padx=6)

        # --- Bottom Bar ---
        bot_frame = ttk.Frame(save_tab)
        bot_frame.pack(fill="x", padx=10, pady=10)

        ttk.Button(
            bot_frame, text="Run Validation Check", command=self.run_validation
        ).pack(side="left")
        ttk.Button(
            bot_frame, text="SAVE TEMPLATE", command=self.save_template, width=20
        ).pack(side="right")
        ttk.Button(
            bot_frame, text="Combine Outputs", command=self.combine_outputs
        ).pack(side="right", padx=6)

        # Populate schema listbox with current schema details
        self._render_schema_list()

    # --- Logic Methods ---

    def _render_schema_list(self) -> None:
        """Update the schema tab listbox with field/synonym info."""
        if not hasattr(self, "schema_listbox"):
            return
        self.schema_listbox.delete(0, tk.END)
        if not self.target_schema:
            self.schema_listbox.insert(tk.END, "No schema fields found.")
            return
        for field, synonyms in self.target_schema.items():
            syn_text = ", ".join(synonyms) if synonyms else "(no synonyms)"
            self.schema_listbox.insert(tk.END, f"{field}: {syn_text}")

    # --- Async helpers ---

    def _set_busy(self, message: str) -> None:
        self._busy = True
        self.status_var.set(message)
        try:
            self.root.config(cursor="wait")
        except Exception:
            pass

    def _clear_busy(self, message: str | None = None) -> None:
        self._busy = False
        if message:
            self.status_var.set(message)
        else:
            self.status_var.set("Ready")
        try:
            self.root.config(cursor="")
        except Exception:
            pass

    def _run_worker(self, task, on_success=None, on_error=None, message: str = "Working...") -> None:
        if self._busy:
            return
        self._set_busy(message)

        def runner():
            try:
                result = task()
            except Exception as exc:
                if on_error:
                    self.root.after(0, lambda: on_error(exc))
                else:
                    self.root.after(
                        0,
                        lambda: messagebox.showerror("Error", str(exc)),
                    )
                self.root.after(0, self._clear_busy)
                return
            if on_success:
                self.root.after(0, lambda: on_success(result))
            else:
                self.root.after(0, self._clear_busy)

        threading.Thread(target=runner, daemon=True).start()

    def _refresh_mapping_dropdowns(self) -> None:
        """Sync mapping dropdowns after schema reload."""
        if hasattr(self, "alias_combo"):
            self.alias_combo["values"] = self.target_fields
        if hasattr(self, "combine_on_combo"):
            self.combine_on_combo["values"] = [""] + self.target_fields

    def _selected_schema_sheet(self) -> Optional[str | int]:
        """Return selected sheet name/index for schema loading."""
        if not self.schema_sheet_listbox.curselection():
            return None
        idx = self.schema_sheet_listbox.curselection()[0]
        try:
            return self.schema_sheet_listbox.get(idx)
        except Exception:
            return None

    def _prompt_sheet_choice(self, sheet_names: List[str]) -> Optional[str]:
        """Ask user to pick a sheet if multiple exist."""
        if not sheet_names:
            return None
        if len(sheet_names) == 1:
            return sheet_names[0]
        choice = simpledialog.askstring(
            "Sheet selection",
            f"Available sheets:\n{', '.join(sheet_names)}\n\nEnter sheet name or index (0-based). Leave blank for first sheet.",
            parent=self.root,
        )
        if choice is None or choice.strip() == "":
            return sheet_names[0]
        choice = choice.strip()
        # Allow numeric index
        if choice.isdigit() and int(choice) < len(sheet_names):
            return sheet_names[int(choice)]
        if choice in sheet_names:
            return choice
        messagebox.showwarning("Sheet not found", f"Could not find sheet '{choice}'. Using first sheet.")
        return sheet_names[0]

    def _set_schema_candidates(self, annotated: List[Dict[str, object]]) -> None:
        """Update UI list with precomputed candidates."""
        self.schema_candidates = annotated
        self.suggestions_listbox.delete(0, tk.END)
        for idx, cand in enumerate(annotated):
            label = cand.get("label", f"Candidate {idx+1}")
            score = cand.get("score")
            suffix = f" (score {score:.2f})" if isinstance(score, (int, float)) else ""
            self.suggestions_listbox.insert(tk.END, f"{label}{suffix}")
        self.suggestion_detail_var.set("")

    def _build_schema_candidates(self, df: pd.DataFrame, headers: List[str], source_path: Path) -> None:
        """Build candidate header sets based on data-type hints and heuristics."""
        data_type = self.schema_data_type_var.get()
        target_fields = list(self.target_schema.keys()) if self.target_schema else []
        annotated = build_schema_candidates(
            df=df,
            headers=headers,
            data_type=data_type,
            target_fields=target_fields,
        )
        self._set_schema_candidates(annotated)

    def _build_schema_candidates_async(self, df: pd.DataFrame, headers: List[str], source_path: Path) -> None:
        """Run candidate building off the UI thread."""

        def work():
            data_type = self.schema_data_type_var.get()
            target_fields = list(self.target_schema.keys()) if self.target_schema else []
            return build_schema_candidates(
                df=df,
                headers=headers,
                data_type=data_type,
                target_fields=target_fields,
            )

        def on_success(result):
            self._set_schema_candidates(result)
            self._clear_busy("Candidates ready")

        def on_error(exc: Exception):
            messagebox.showerror("Suggestion error", str(exc))
            self._clear_busy("Error")

        self._run_worker(work, on_success=on_success, on_error=on_error, message="Building header suggestions...")

    def _apply_candidate_schema(self) -> None:
        """Apply selected candidate headers to schema and final preview."""
        if not self.schema_candidates or not self.suggestions_listbox.curselection():
            messagebox.showwarning("No candidate", "Select a suggested header set first.")
            return
        idx = self.suggestions_listbox.curselection()[0]
        candidate = self.schema_candidates[idx]
        headers = candidate.get("headers", [])
        if not headers:
            messagebox.showwarning("Empty candidate", "Selected candidate has no headers.")
            return

        # Rebuild a preview DF with these headers if possible
        if self.schema_preview_tree.get_children():
            # Reconstruct a tiny DF from preview grid
            cols = self.schema_preview_tree["columns"]
            rows = []
            for item in self.schema_preview_tree.get_children():
                rows.append(self.schema_preview_tree.item(item)["values"])
            df_preview = pd.DataFrame(rows, columns=cols)
            # Apply new headers length alignment
            adjusted_headers = list(headers)
            if len(adjusted_headers) < len(df_preview.columns):
                adjusted_headers.extend(df_preview.columns[len(adjusted_headers):])
            elif len(adjusted_headers) > len(df_preview.columns):
                adjusted_headers = adjusted_headers[: len(df_preview.columns)]
            df_preview.columns = adjusted_headers
            self._apply_headers_to_schema(
                adjusted_headers,
                Path(self.schema_source_path or "schema.json"),
                df_preview,
                prev_fields=self.target_fields,
            )
        else:
            self._apply_headers_to_schema(headers, Path(self.schema_source_path or "schema.json"), prev_fields=self.target_fields)

    def _on_candidate_select(self) -> None:
        """Show details for the selected candidate."""
        if not self.schema_candidates or not self.suggestions_listbox.curselection():
            self.suggestion_detail_var.set("")
            return
        idx = self.suggestions_listbox.curselection()[0]
        cand = self.schema_candidates[idx]
        note = cand.get("note", "")
        score = cand.get("score")
        score_txt = f"Score: {score:.2f}. " if isinstance(score, (int, float)) else ""
        missing = cand.get("missing") or []
        extra = cand.get("extra") or []
        missing_txt = f" Missing: {', '.join(missing[:5])}" + ("..." if len(missing) > 5 else "") if missing else ""
        extra_txt = f" Extra: {', '.join(extra[:5])}" + ("..." if len(extra) > 5 else "") if extra else ""
        self.suggestion_detail_var.set(f"{score_txt}{note}{missing_txt}{extra_txt}")

    def _attach_tooltip(self, widget: tk.Widget, text: str) -> None:
        """Attach a simple tooltip to a widget."""
        tooltip = tk.Toplevel(widget)
        tooltip.withdraw()
        tooltip.overrideredirect(True)
        label = tk.Label(
            tooltip,
            text=text,
            background="#ffffe0",
            relief="solid",
            borderwidth=1,
            justify="left",
        )
        label.pack()

        def enter(_event):
            x = widget.winfo_rootx() + 20
            y = widget.winfo_rooty() + 20
            tooltip.geometry(f"+{x}+{y}")
            tooltip.deiconify()

        def leave(_event):
            tooltip.withdraw()

        widget.bind("<Enter>", enter)
        widget.bind("<Leave>", leave)

    def _apply_and_save_schema(self) -> None:
        """Apply current headers/mapping and immediately save the schema."""
        if not self.target_fields:
            messagebox.showwarning("No schema", "Load or build a schema first.")
            return
        self._save_schema_file()

    def _reset_mapping_to_defaults(self) -> None:
        """Reset headers to last saved snapshot or built-in defaults."""
        if self.saved_schema_snapshot:
            headers = list(self.saved_schema_snapshot.keys())
            self._apply_headers_to_schema(
                headers, self.current_schema_path or Path("schema.json"), prev_fields=self.target_fields
            )
            messagebox.showinfo("Reset", "Schema reset to last saved version.")
        else:
            self._reset_to_builtin_schema()

    # --- Custom mapping editor ---
    def _open_custom_map_editor(self) -> None:
        """Allow users to edit/save/load custom header mappings (text -> replacement)."""
        top = tk.Toplevel(self.root)
        top.title("Custom Header Mapping")
        top.geometry("700x420")

        path_var = tk.StringVar(value=str(self.custom_map_path))

        path_row = ttk.Frame(top)
        path_row.pack(fill="x", padx=8, pady=6)
        ttk.Label(path_row, text="Mapping file:").pack(side="left")
        ttk.Entry(path_row, textvariable=path_var, width=50).pack(side="left", padx=6)
        ttk.Button(path_row, text="Load", command=lambda: self._load_map_from_file(path_var, text_widget)).pack(side="left", padx=4)
        ttk.Button(path_row, text="Save As", command=lambda: self._save_map_to_file(path_var, text_widget, save_as=True)).pack(side="left", padx=4)
        ttk.Button(path_row, text="Apply to current headers", command=lambda: self._apply_map_to_schema(text_widget)).pack(side="right", padx=4)

        ttk.Label(
            top,
            text="One mapping per line. Format: old -> new (or old:new). Lines starting with # are ignored.",
            foreground="gray",
        ).pack(anchor="w", padx=8, pady=(0, 4))

        text_widget = tk.Text(top, wrap="none")
        text_widget.pack(fill="both", expand=True, padx=8, pady=4)

        # Seed with existing file if present
        loaded = self._load_map_from_file(path_var, text_widget, silent=True)
        if not loaded:
            example = (
                "# Example mappings:\n"
                "# prod code -> product_id\n"
                "# tammikuu -> january\n"
                "# amount € -> sales_amount\n"
            )
            text_widget.insert(tk.END, example)

    def _parse_header_map_text(self, raw: str) -> Dict[str, str]:
        mapping: Dict[str, str] = {}
        for line in raw.splitlines():
            stripped = line.strip()
            if not stripped or stripped.startswith("#"):
                continue
            sep = "->" if "->" in stripped else (":" if ":" in stripped else None)
            if not sep:
                continue
            parts = stripped.split(sep, 1)
            if len(parts) != 2:
                continue
            src = parts[0].strip()
            dst = parts[1].strip()
            if src and dst:
                mapping[src.lower()] = dst
        return mapping

    def _load_map_from_file(self, path_var: tk.StringVar, text_widget: tk.Text, silent: bool = False) -> bool:
        """Load mapping file content into text widget. Returns True if loaded."""
        path = Path(path_var.get())
        try:
            if path.exists():
                content = path.read_text(encoding="utf-8")
                text_widget.delete("1.0", tk.END)
                text_widget.insert(tk.END, content)
                self.custom_map_path = path
                self.custom_map_label_var.set(str(path))
                return True
            elif not silent:
                messagebox.showinfo("Not found", f"{path} does not exist. Start typing to create it.")
        except Exception as exc:
            if not silent:
                messagebox.showerror("Load failed", str(exc))
        return False

    def _save_map_to_file(self, path_var: tk.StringVar, text_widget: tk.Text, save_as: bool = False) -> None:
        """Save mapping text to a file; allow choosing a new name."""
        path = Path(path_var.get())
        if save_as:
            chosen = filedialog.asksaveasfilename(
                defaultextension=".txt",
                filetypes=[("Text", "*.txt"), ("All files", "*.*")],
                initialdir=path.parent if path else SCHEMA_DIR,
                initialfile=path.name if path else "header_map.txt",
                title="Save mapping as",
            )
            if not chosen:
                return
            path = Path(chosen)
            path_var.set(str(path))
        try:
            path.parent.mkdir(parents=True, exist_ok=True)
            text = text_widget.get("1.0", tk.END)
            path.write_text(text, encoding="utf-8")
            self.custom_map_path = path
            self.custom_map_label_var.set(str(path))
            messagebox.showinfo("Saved", f"Mapping saved to {path}")
        except Exception as exc:
            messagebox.showerror("Save failed", str(exc))

    def _apply_map_to_schema(self, text_widget: tk.Text) -> None:
        """Apply custom mapping to current headers and refresh schema."""
        if not self.target_fields:
            messagebox.showwarning("No schema", "Load or build a schema first.")
            return
        mapping = self._parse_header_map_text(text_widget.get("1.0", tk.END))
        if not mapping:
            messagebox.showwarning("Empty mapping", "Add mappings in the editor first.")
            return

        new_headers: List[str] = []
        for h in self.target_fields:
            replacement = mapping.get(str(h).lower(), h)
            new_headers.append(replacement)

        # Reuse existing preview DF if available for final preview
        df_preview = None
        if self.final_preview_tree.get_children():
            cols = self.final_preview_tree["columns"]
            rows = []
            for item in self.final_preview_tree.get_children():
                rows.append(self.final_preview_tree.item(item)["values"])
            df_preview = pd.DataFrame(rows, columns=cols)
            if len(new_headers) != len(df_preview.columns):
                # align lengths
                adjusted_headers = list(new_headers)
                if len(adjusted_headers) < len(df_preview.columns):
                    adjusted_headers.extend(df_preview.columns[len(adjusted_headers):])
                else:
                    adjusted_headers = adjusted_headers[: len(df_preview.columns)]
                df_preview.columns = adjusted_headers
            else:
                df_preview.columns = new_headers

        self._apply_headers_to_schema(
            new_headers,
            self.current_schema_path or Path("schema.json"),
            df_preview,
            suppress_msg=True,
            prev_fields=self.target_fields,
        )
        messagebox.showinfo("Mapping applied", "Custom mapping applied to schema headers.")

    def _guess_schema_path(self) -> Path:
        for candidate in DEFAULT_SCHEMA_CANDIDATES:
            if Path(candidate).exists():
                return Path(candidate)
        return DEFAULT_SCHEMA_CANDIDATES[0]

    def _reload_target_schema(self, path: Optional[Path] | None = None) -> None:
        """Reload schema from a chosen file or default location."""
        try:
            schema = load_target_schema(path)
        except Exception as exc:
            messagebox.showerror("Schema load failed", str(exc))
            return

        self.target_schema = schema
        self.target_fields = list(schema.keys())
        self.current_schema_path = Path(path) if path else self._guess_schema_path()
        self.schema_path_var.set(str(self.current_schema_path))
        self.schema_fields_var.set(", ".join(self.target_fields) if self.target_fields else "-")
        self._render_schema_list()
        self._refresh_mapping_dropdowns()
        self._update_diff_labels(self.target_fields, self.target_fields)

    def _render_schema_list(self) -> None:
        """Update the schema tab listbox with field/synonym info."""
        if not hasattr(self, "schema_listbox"):
            return
        self.schema_listbox.delete(0, tk.END)
        if not self.target_schema:
            self.schema_listbox.insert(tk.END, "No schema fields found.")
            return
        for field, synonyms in self.target_schema.items():
            syn_text = ", ".join(synonyms) if synonyms else "(no synonyms)"
            self.schema_listbox.insert(tk.END, f"{field}: {syn_text}")

    def _reset_to_builtin_schema(self) -> None:
        """Reset schema to built-in defaults, ignoring files/config."""
        headers = list(TARGET_SCHEMA.keys())
        self._apply_headers_to_schema(
            headers,
            Path("schema.json"),
            df=None,
            suppress_msg=False,
            prev_fields=self.target_fields,
        )

    def _open_schema_folder(self) -> None:
        """Open the schema directory so users can drop/update files."""
        try:
            SCHEMA_DIR.mkdir(parents=True, exist_ok=True)
            os.startfile(str(SCHEMA_DIR.resolve()))
        except Exception as exc:
            messagebox.showinfo(
                "Schema folder",
                f"Schema folder is at {SCHEMA_DIR.resolve()}\n\n{exc}",
            )

    def _choose_schema_file(self) -> None:
        """Pick a schema file and reload mapping options."""
        path = filedialog.askopenfilename(
            filetypes=[("Schema JSON", "*.json"), ("All files", "*.*")],
            initialdir=SCHEMA_DIR,
        )
        if not path:
            return
        self._reload_target_schema(Path(path))

    def _reset_schema_preview(self) -> None:
        """Clear schema preview and selections (keeps existing schema fields)."""
        self.schema_source_path = None
        self.schema_sheet_listbox.delete(0, tk.END)
        self.schema_preview_tree.delete(*self.schema_preview_tree.get_children())
        self.schema_preview_tree["columns"] = ()
        self.final_preview_tree.delete(*self.final_preview_tree.get_children())
        self.final_preview_tree["columns"] = ()
        self.suggestions_listbox.delete(0, tk.END)
        self.schema_candidates = []
        self.suggestion_detail_var.set("")
        self.schema_header_row_var.set("0")
        self.schema_skiprows_var.set("")
        self.schema_delimiter_var.set(",")
        self.schema_encoding_var.set("utf-8")

    def _reload_schema_preview(self, update_schema: bool = False, suppress_msg: bool = False, use_heuristics: bool = False) -> None:
        """Reload preview for schema source and optionally update schema fields."""
        if not self.schema_source_path:
            if not suppress_msg:
                messagebox.showwarning("No file", "Load a schema file first.")
            return

        path = Path(self.schema_source_path)
        is_csv = path.suffix.lower() == ".csv"
        try:
            header_row = int(self.schema_header_row_var.get() or 0)
        except ValueError:
            header_row = 0
        skiprows = parse_skiprows(self.schema_skiprows_var.get())
        delimiter = self.schema_delimiter_var.get() or ","
        encoding = self.schema_encoding_var.get() or "utf-8"

        sheet = None
        if not is_csv:
            selection = self._selected_schema_sheet()
            sheet = selection if selection is not None else 0

        try:
            preview_df = read_preview_frame(
                path=path,
                source_type="csv" if is_csv else "excel",
                sheet=sheet,
                header_row=header_row,
                skiprows=skiprows,
                nrows=DEFAULT_PREVIEW_ROWS,
                delimiter=delimiter,
                encoding=encoding,
            )
            if not is_csv:
                try:
                    normalized_headers, _merged = get_normalized_headers(
                        path=path, sheet=sheet or 0, header_row=header_row, skiprows=skiprows
                    )
                    preview_df = apply_normalized_headers(preview_df, normalized_headers)
                except Exception:
                    pass
        except Exception as exc:
            if not suppress_msg:
                messagebox.showerror("Preview failed", str(exc))
            return

        # Update preview grid
        self.schema_preview_tree.delete(*self.schema_preview_tree.get_children())
        self.schema_preview_tree["columns"] = list(preview_df.columns)
        for col in preview_df.columns:
            self.schema_preview_tree.heading(col, text=str(col))
            self.schema_preview_tree.column(col, width=120)
        for row in preview_df.itertuples(index=False):
            self.schema_preview_tree.insert("", tk.END, values=list(row))

        if update_schema or use_heuristics:
            headers = [str(col) for col in preview_df.columns if str(col).strip()]
            if headers:
                if use_heuristics:
                    self._build_schema_candidates_async(preview_df, headers, path)
                else:
                    self._apply_headers_to_schema(headers, path, preview_df, suppress_msg)
            elif not suppress_msg:
                messagebox.showwarning("No headers", "Could not detect headers to build schema.")

    def _save_schema_file(self) -> None:
        """Persist current schema fields/synonyms to a JSON file."""
        if not self.target_schema:
            messagebox.showwarning("No schema to save", "Load or build a schema first.")
            return

        # Default to data/schemas/schema.json unless we already have a JSON schema path
        default_path = (
            self.current_schema_path
            if self.current_schema_path and self.current_schema_path.suffix.lower() == ".json"
            else SCHEMA_DIR / "schema.json"
        )
        path = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("Schema JSON", "*.json"), ("All files", "*.*")],
            initialdir=default_path.parent if default_path else SCHEMA_DIR,
            initialfile=default_path.name if default_path else "schema.json",
            title="Save schema as",
        )
        if not path:
            return

        target_path = Path(path)
        if target_path.suffix.lower() != ".json":
            target_path = target_path.with_suffix(".json")

        payload = {str(k): list(v) if isinstance(v, list) else [] for k, v in self.target_schema.items()}
        try:
            target_path.parent.mkdir(parents=True, exist_ok=True)
            with open(target_path, "w", encoding="utf-8") as handle:
                json.dump(payload, handle, indent=2)
            self.current_schema_path = target_path
            self.schema_path_var.set(str(self.current_schema_path))
            self.saved_schema_snapshot = {k: list(v) for k, v in self.target_schema.items()}
            messagebox.showinfo("Schema saved", f"Saved schema to {target_path}")
        except Exception as exc:
            messagebox.showerror("Save failed", str(exc))

    def _apply_headers_to_schema(
        self,
        headers: List[str],
        source_path: Path,
        df: Optional[pd.DataFrame] = None,
        suppress_msg: bool = False,
        prev_fields: Optional[List[str]] = None,
    ) -> None:
        """Apply a header list to the current schema state and refresh previews."""
        prev_fields = prev_fields or self.target_fields
        schema = {h: [] for h in headers}
        self.target_schema = schema
        self.target_fields = headers
        self.current_schema_path = source_path
        self.schema_path_var.set(str(self.current_schema_path))
        self.schema_fields_var.set(", ".join(headers))
        self._render_schema_list()
        self._refresh_mapping_dropdowns()
        self._update_diff_labels(prev_fields, headers)
        if df is not None:
            self._update_final_preview(df)
        if not suppress_msg:
            messagebox.showinfo(
                "Schema updated",
                f"Updated schema with {len(headers)} fields from {source_path.name}.",
            )

    def _update_final_preview(self, df: pd.DataFrame) -> None:
        """Render a final preview grid with current headers."""
        self.final_preview_tree.delete(*self.final_preview_tree.get_children())
        self.final_preview_tree["columns"] = list(df.columns)
        for col in df.columns:
            if col in self.final_extra_set:
                heading_style = "ExtraHeading.Treeview.Heading"
            elif col in self.final_missing_set:
                heading_style = "MissingHeading.Treeview.Heading"
            else:
                heading_style = None
            self._set_heading(self.final_preview_tree, col, str(col), heading_style)
            self.final_preview_tree.column(col, width=120)
        for row in df.itertuples(index=False):
            self.final_preview_tree.insert("", tk.END, values=list(row))

    def _update_diff_labels(self, prev_fields: List[str], new_headers: List[str]) -> None:
        """Update missing/extra labels comparing previous fields to new headers."""
        prev_set = set(prev_fields)
        new_set = set(new_headers)
        missing = sorted(list(prev_set - new_set))
        extra = sorted(list(new_set - prev_set))
        self.final_missing_set = set(missing)
        self.final_extra_set = set(extra)
        missing_txt = ", ".join(missing[:5]) + ("..." if len(missing) > 5 else "") if missing else "-"
        extra_txt = ", ".join(extra[:5]) + ("..." if len(extra) > 5 else "") if extra else "-"
        self.final_diff_missing_var.set(missing_txt)
        self.final_diff_extra_var.set(extra_txt)

    def _init_styles(self) -> None:
        """Configure Treeview heading styles for extra/missing columns."""
        style = ttk.Style(self.root)
        style.configure("ExtraHeading.Treeview.Heading", foreground="blue")
        style.configure("MissingHeading.Treeview.Heading", foreground="red")
        if hasattr(self, "preview_tree"):
            self.preview_tree.tag_configure("metadata", background="#fff2cc")

    def _set_heading(self, tree: ttk.Treeview, col: str, text: str, style: Optional[str]) -> None:
        """Set heading with optional style; tolerate Tk versions without style support."""
        try:
            if style:
                tree.heading(col, text=text, style=style)
            else:
                tree.heading(col, text=text)
        except tk.TclError:
            tree.heading(col, text=text)

    def _load_schema_from_excel(self) -> None:
        """Read headers from a sample Excel/CSV file and treat them as the target schema."""
        path = filedialog.askopenfilename(
            filetypes=[("Excel/CSV", "*.xlsx *.xls *.csv")],
            initialdir=SCHEMA_DIR,
            title="Select example output file",
        )
        if not path:
            return
        if not path.lower().endswith((".xlsx", ".xls", ".csv")):
            messagebox.showwarning("Unsupported file", "Select an Excel or CSV file for schema loading.")
            return

        self.schema_source_path = path
        is_csv = path.lower().endswith(".csv")
        self.schema_delimiter_var.set("," if is_csv else self.delimiter_var.get())
        self.schema_encoding_var.set("utf-8")
        self.schema_skiprows_var.set("")

        if is_csv:
            self.schema_sheet_listbox.delete(0, tk.END)
            self.schema_sheet_listbox.insert(tk.END, "CSV")
            self.schema_sheet_listbox.selection_set(0)
        else:
            try:
                sheets = sheet_names(Path(path))
                self.schema_sheet_listbox.delete(0, tk.END)
                for name in sheets:
                    self.schema_sheet_listbox.insert(tk.END, name)
                if sheets:
                    self.schema_sheet_listbox.selection_set(0)
            except Exception as exc:
                messagebox.showerror("Failed to open workbook", str(exc))
                return

        # Guess header row using preview data
        try:
            sheet_for_preview = (
                self.schema_sheet_listbox.get(0) if (not is_csv and self.schema_sheet_listbox.size()) else None
            )
            preview = read_preview_frame(
                Path(path),
                source_type="csv" if is_csv else "excel",
                sheet=sheet_for_preview,
                header_row=None,
                skiprows=[],
                nrows=DEFAULT_PREVIEW_ROWS,
                delimiter="," if is_csv else self.delimiter_var.get(),
                encoding="utf-8",
            )
            guessed_header = guess_header_row(preview)
            self.schema_header_row_var.set(str(guessed_header))
        except Exception:
            self.schema_header_row_var.set("0")

        self._reload_schema_preview(update_schema=True, suppress_msg=True)
        messagebox.showinfo("Schema loaded", "Schema fields updated from the selected file.")

    def select_file(self) -> None:
        path = filedialog.askopenfilename(filetypes=[("Data", "*.xlsx *.xls *.csv")])
        if not path:
            return

        self.file_path = path
        self.file_label.config(text=os.path.basename(path))
        self.source_type = "csv" if path.lower().endswith(".csv") else "excel"

        # Reset
        self.mapping.clear()
        self.header_cells.clear()
        self.columns.clear()
        self.columns_listbox.delete(0, tk.END)
        self.mapping_tree.delete(*self.mapping_tree.get_children())
        self.combine_sheets_var.set(False)

        # Check for existing template in the same folder
        potential_tpl = default_template_path(Path(path))
        if potential_tpl.exists():
            if messagebox.askyesno(
                "Template Found",
                "An existing template was found for this file. Load it?",
            ):
                if self.source_type == "excel":
                    try:
                        self.sheet_names = sheet_names(Path(path))
                        self.sheet_listbox.config(state="normal")
                        self.sheet_listbox.delete(0, tk.END)
                        for name in self.sheet_names:
                            self.sheet_listbox.insert(tk.END, name)
                    except Exception:
                        # Fall back to treating as CSV if Excel engine fails
                        self.source_type = "csv"
                self._load_from_file(potential_tpl)
                return

        # Initialize Sheet selection
        if self.source_type == "csv":
            self.sheet_names = ["CSV"]
            self.sheet_listbox.delete(0, tk.END)
            self.sheet_listbox.insert(tk.END, "CSV")
            self.sheet_listbox.selection_set(0)
            self.sheet_listbox.config(state="disabled")
            self._auto_suggest_header_row()
            self.load_headers()
        else:
            try:
                self.sheet_names = sheet_names(Path(path))
                self.sheet_listbox.config(state="normal")
                self.sheet_listbox.delete(0, tk.END)
                for name in self.sheet_names:
                    self.sheet_listbox.insert(tk.END, name)
                if self.sheet_names:
                    self.sheet_listbox.selection_set(0)
                self._auto_suggest_header_row()
                self.load_headers()
            except Exception:
                # Treat as CSV fallback
                self.source_type = "csv"
                self.sheet_names = ["CSV"]
                self.sheet_listbox.delete(0, tk.END)
                self.sheet_listbox.insert(tk.END, "CSV")
                self.sheet_listbox.selection_set(0)
                self.sheet_listbox.config(state="disabled")
                self._auto_suggest_header_row()
                self.load_headers()

    def _auto_suggest_header_row(self) -> None:
        """Read raw rows to guess where the header is."""
        try:
            sheet = None
            if self.source_type != "csv":
                sheets = self._selected_sheets()
                sheet = sheets[0] if sheets else 0
            df = read_preview_frame(
                Path(self.file_path),
                source_type=self.source_type,
                sheet=sheet,
                header_row=None,
                skiprows=[],
                nrows=15,
                delimiter=self.delimiter_var.get(),
                encoding=self.encoding_var.get(),
            )
            suggested = guess_header_row(df)
            self.header_row_var.set(str(suggested))
        except:
            pass

    def load_headers(self) -> None:
        if self.source_type == "sql":
            return
        opts = self._get_options()
        if not opts:
            return
        self._clear_metadata_selection()

        def work():
            headers: List[str] = []
            preview_frames: List[pd.DataFrame] = []
            warnings: List[str] = []
            path = Path(self.file_path)

            if self.source_type == "excel":
                sheet_targets = opts["sheets"] if opts["combine_sheets"] else [opts["sheet"]]
                per_sheet_headers: List[List[str]] = []
                for sheet in sheet_targets:
                    hdrs, _ = get_normalized_headers(
                        path,
                        sheet,
                        opts["header"],
                        opts["skiprows"],
                    )
                    per_sheet_headers.append(hdrs)
                if per_sheet_headers:
                    ref = per_sheet_headers[0]
                    common = [h for h in ref if all(h in hs for hs in per_sheet_headers[1:])]
                    headers = common if opts["combine_sheets"] else ref
                    if opts["combine_sheets"] and len(per_sheet_headers) > 1:
                        diffs = []
                        base_set = set(ref)
                        for idx, hs in enumerate(per_sheet_headers[1:], start=2):
                            extra = set(hs) - base_set
                            missing = base_set - set(hs)
                            if extra or missing:
                                diffs.append(f"Sheet {idx}: +{len(extra)} / -{len(missing)}")
                        if diffs:
                            warnings.append(
                                "Selected sheets have different columns. Using common columns only. "
                                + "; ".join(diffs)
                            )

                # Preview frames
                sheets = sheet_targets
                for sheet in sheets:
                    df = read_preview_frame(
                        path,
                        source_type="excel",
                        sheet=sheet,
                        header_row=opts["header"],
                        skiprows=opts["skiprows"],
                        nrows=DEFAULT_PREVIEW_ROWS,
                    )
                    if headers:
                        if opts["combine_sheets"]:
                            for col in headers:
                                if col not in df.columns:
                                    df[col] = pd.NA
                            df = df[headers]
                        else:
                            df = apply_normalized_headers(df, headers)
                    if opts["combine_sheets"]:
                        df["source_sheet"] = str(sheet)
                    preview_frames.append(df)
            else:
                df = read_preview_frame(
                    path,
                    source_type="csv",
                    sheet=None,
                    header_row=opts["header"],
                    skiprows=opts["skiprows"],
                    nrows=DEFAULT_PREVIEW_ROWS,
                    delimiter=opts["sep"],
                    encoding=opts["encoding"],
                )
                headers = list(map(str, df.columns))
                preview_frames.append(df)

            preview_df = (
                pd.concat(preview_frames, ignore_index=True) if len(preview_frames) > 1 else preview_frames[0]
            )
            return {"headers": headers, "preview_df": preview_df, "warnings": warnings}

        def on_success(result: dict):
            headers = result.get("headers", [])
            preview_df = result.get("preview_df")
            warnings = result.get("warnings", [])

            self.columns = headers
            self.columns_listbox.delete(0, tk.END)
            for c in self.columns:
                self.columns_listbox.insert(tk.END, c)

            if preview_df is not None and isinstance(preview_df, pd.DataFrame):
                self.preview_df = preview_df
                self.preview_tree.delete(*self.preview_tree.get_children())
                self.preview_tree["columns"] = list(preview_df.columns)
                for col in preview_df.columns:
                    self.preview_tree.heading(col, text=col)
                    self.preview_tree.column(col, width=100)
                for row in preview_df.itertuples(index=False):
                    self.preview_tree.insert("", tk.END, values=list(row))

            if warnings:
                messagebox.showwarning("Sheet mismatch", "\n".join(warnings))
            self._update_info_panel()
            self._clear_busy("Preview loaded")

        def on_error(exc: Exception):
            messagebox.showerror("Read Error", f"Could not read headers: {exc}")
            self._clear_busy("Error")

        self._run_worker(work, on_success=on_success, on_error=on_error, message="Loading preview...")

    def load_preview(self) -> None:
        if self.source_type == "sql":
            return
        opts = self._get_options()
        if not opts:
            return
        self._clear_metadata_selection()

        try:
            frames: List[pd.DataFrame] = []
            if self.source_type == "csv":
                df = read_preview_frame(
                    Path(self.file_path),
                    source_type="csv",
                    sheet=None,
                    header_row=opts["header"],
                    skiprows=opts["skiprows"],
                    nrows=DEFAULT_PREVIEW_ROWS,
                    delimiter=opts["sep"],
                    encoding=opts["encoding"],
                )
                frames.append(df)
            else:
                sheets = opts["sheets"] if opts["combine_sheets"] else [opts["sheet"]]
                for sheet in sheets:
                    df = read_preview_frame(
                        Path(self.file_path),
                        source_type="excel",
                        sheet=sheet,
                        header_row=opts["header"],
                        skiprows=opts["skiprows"],
                        nrows=DEFAULT_PREVIEW_ROWS,
                    )
                    if self.columns:
                        if opts["combine_sheets"]:
                            # Align to common columns for multi-sheet preview
                            for col in self.columns:
                                if col not in df.columns:
                                    df[col] = pd.NA
                            df = df[self.columns]
                        else:
                            df = apply_normalized_headers(df, self.columns)
                    if opts["combine_sheets"]:
                        df["source_sheet"] = str(sheet)
                    frames.append(df)

            if not frames:
                return

            df = pd.concat(frames, ignore_index=True) if len(frames) > 1 else frames[0]
            self.preview_df = df

            # Update Treeview
            self.preview_tree.delete(*self.preview_tree.get_children())
            self.preview_tree["columns"] = list(df.columns)
            for col in df.columns:
                self.preview_tree.heading(col, text=col)
                self.preview_tree.column(col, width=100)

            for row in df.itertuples(index=False):
                self.preview_tree.insert("", tk.END, values=list(row))
        except:
            pass

    def add_mapping(self) -> None:
        sel = self.columns_listbox.curselection()
        target = self.alias_var.get()
        if not sel or not target:
            return

        src = self.columns_listbox.get(sel[0])
        self.mapping[src] = target

        # Store metadata
        opts = self._get_options()
        col_idx = self.columns.index(src) if src in self.columns else -1
        self.header_cells[src] = HeaderCell(src, col_idx, opts["header"], target)

        self._refresh_mapping_view()

    def apply_smart_mapping(self) -> None:
        """Use heuristics to guess mappings."""
        if not self.columns:
            return
        suggestions = auto_map_columns(self.columns, self.target_schema)

        opts = self._get_options()
        for src, target in suggestions.items():
            self.mapping[src] = target
            col_idx = self.columns.index(src)
            self.header_cells[src] = HeaderCell(src, col_idx, opts["header"], target)

        self._refresh_mapping_view()
        messagebox.showinfo(
            "Auto-Map", f"Suggested {len(suggestions)} mappings based on schema."
        )

    def _refresh_mapping_view(self) -> None:
        self.mapping_tree.delete(*self.mapping_tree.get_children())
        is_unpivot = self.unpivot_var.get()

        for src, target in self.mapping.items():
            # If unpivoting, mapped columns are 'Identifiers'. Others are 'Values'.
            role = "Identifier" if is_unpivot else "Standard"
            self.mapping_tree.insert("", tk.END, values=(src, target, role))

    def remove_mapping(self) -> None:
        sel = self.mapping_tree.selection()
        if not sel:
            return
        src = self.mapping_tree.item(sel[0])["values"][0]
        if src in self.mapping:
            del self.mapping[src]
            del self.header_cells[src]
            self._refresh_mapping_view()

    def run_validation(self) -> None:
        # Simple check against schema
        missing = [t for t in self.target_schema if t not in self.mapping.values()]

        msg = "Validation Report:\n"
        if not missing:
            msg += "- All target fields are mapped.\n"
        else:
            msg += f"! Missing fields: {', '.join(missing)}\n"

        if self.unpivot_var.get():
            msg += "Unpivot Mode is ON.\n"

        messagebox.showinfo("Validation", msg)

    def save_template(self) -> None:
        if not self.file_path or not self.mapping:
            return

        opts = self._get_options()
        id_cols = list(self.mapping.values()) if self.unpivot_var.get() else []
        combine_keys = self._parse_combine_on()
        dedupe_keys = self._parse_dedupe_on()
        drop_null_threshold = self._parse_float(self.drop_null_threshold_var.get())

        all_headers = list(self.header_cells.values()) + list(self.metadata_cells)
        tpl = Template(
            source_file=os.path.basename(self.file_path),
            source_type=self.source_type,
            sheet=opts["sheet"],
            sheets=opts["sheets"],
            header_row=opts["header"],
            skiprows=opts["skiprows"],
            delimiter=opts["sep"],
            encoding=opts["encoding"],
            columns=list(self.mapping.keys()),
            column_mappings=self.mapping,
            headers=all_headers,
            combine_sheets=opts["combine_sheets"],
            combine_on=combine_keys,
            connection_name=self.connection_name_var.get() or None,
            sql_table=opts["sql_table"],
            sql_query=opts["sql_query"],
            trim_strings=self.trim_strings_var.get(),
            drop_empty_rows=self.drop_empty_rows_var.get(),
            drop_null_columns_threshold=drop_null_threshold,
            dedupe_on=dedupe_keys,
            strip_thousands=self.strip_thousands_var.get(),
            unpivot=self.unpivot_var.get(),
            id_columns=id_cols,
            var_name=self.var_name_var.get(),
            value_name=self.val_name_var.get(),
            output_dir=os.path.dirname(self.file_path) if self.file_path else str(Path("data/output")),
        )

        out_path = default_template_path(Path(self.file_path))
        save_template(tpl, out_path)
        added, cfg_path = learn_synonyms_from_mapping(self.mapping)
        extra = (
            f"\nCaptured {added} new header pattern(s) in {cfg_path.name}."
            if added
            else ""
        )
        messagebox.showinfo("Success", f"Template saved to:\n{out_path}{extra}")

    def _get_options(self) -> dict:
        # Helper to safely get parsing options
        try:
            selected_sheets = self._selected_sheets()
            sheet_value = selected_sheets[0] if selected_sheets else 0
            return {
                "header": int(self.header_row_var.get()),
                "skiprows": parse_skiprows(self.skiprows_var.get()),
                "sep": self.delimiter_var.get(),
                "encoding": self.encoding_var.get(),
                "sheet": sheet_value if self.source_type == "excel" else None,
                "sheets": selected_sheets if self.source_type == "excel" else [],
                "combine_sheets": self.combine_sheets_var.get(),
                "sql_table": self.sql_table_var.get() if self.source_type == "sql" else None,
                "sql_query": self.sql_query_var.get() if self.source_type == "sql" else None,
            }
        except ValueError:
            return {
                "header": 0,
                "skiprows": [],
                "sep": ",",
                "encoding": "utf-8",
                "sheet": 0,
                "sheets": [],
                "combine_sheets": False,
                "sql_table": None,
                "sql_query": None,
            }

    def _selected_sheets(self) -> List[str]:
        if self.source_type != "excel":
            return []
        selections = self.sheet_listbox.curselection()
        if not selections and self.sheet_listbox.size():
            return [self.sheet_listbox.get(0)]
        return [self.sheet_listbox.get(i) for i in selections]

    def _sheet_stats(self, path: Path, sheets: List[str], header: int, skiprows: List[int]) -> tuple[int, int]:
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception:
            return (0, 0)

        targets = sheets or wb.sheetnames[:1]
        total_rows = 0
        max_cols = 0
        for sheet in targets:
            try:
                ws = wb[sheet] if isinstance(sheet, str) else wb.worksheets[int(sheet)]
            except Exception:
                continue
            total_rows += ws.max_row or 0
            max_cols = max(max_cols, ws.max_column or 0)
        # Rough adjustment for header/skip rows
        effective_rows = max(0, total_rows - header - len(skiprows) - 1)
        return (effective_rows, max_cols)

    def _csv_stats(self, path: Path) -> tuple[int, int]:
        try:
            with path.open(
                "r",
                encoding=self.encoding_var.get() or "utf-8",
                errors="ignore",
            ) as handle:
                first_line = handle.readline()
                line_count = 0
                for _ in handle:
                    line_count += 1
            cols = len(first_line.split(self.delimiter_var.get() or ",")) if first_line else 0
            return (line_count, cols)
        except Exception:
            return (0, 0)

    def _update_info_panel(self) -> None:
        if not self.file_path:
            if self.source_type == "sql" and self.preview_df is not None:
                rows, cols = self.preview_df.shape
                self.info_vars["sheets"].set("Source: SQL")
                self.info_vars["rows"].set(f"Rows: {rows}")
                self.info_vars["cols"].set(f"Columns: {cols}")
            else:
                self.info_vars["sheets"].set("Sheets: -")
                self.info_vars["rows"].set("Rows: -")
                self.info_vars["cols"].set("Columns: -")
            return

        opts = self._get_options()
        if self.source_type == "excel":
            sheet_list = opts.get("sheets") or ([] if opts.get("sheet") is None else [opts.get("sheet")])
            rows, cols = self._sheet_stats(Path(self.file_path), sheet_list, opts["header"], opts["skiprows"])
            sheet_label = ", ".join(map(str, sheet_list)) if sheet_list else "-"
        else:
            rows, cols = self._csv_stats(Path(self.file_path))
            sheet_label = "CSV"

        self.info_vars["sheets"].set(f"Sheets: {sheet_label}")
        self.info_vars["rows"].set(f"Rows: {rows}")
        self.info_vars["cols"].set(f"Columns: {cols}")

    def reset_view(self) -> None:
        self.file_path = None
        self.file_label.config(text="No file selected")
        self.sheet_names = []
        self.columns = []
        self.mapping.clear()
        self.header_cells.clear()
        self.preview_df = None
        self.sheet_listbox.delete(0, tk.END)
        self.columns_listbox.delete(0, tk.END)
        self.mapping_tree.delete(*self.mapping_tree.get_children())
        self.combine_sheets_var.set(False)
        self.unpivot_var.set(False)
        self.combine_on_var.set("")
        self.connection_name_var.set("")
        self.trim_strings_var.set(True)
        self.drop_empty_rows_var.set(False)
        self.drop_null_threshold_var.set("")
        self.dedupe_on_var.set("")
        self.strip_thousands_var.set(False)
        self.sql_table_var.set("")
        self.sql_query_var.set("")
        self.alias_var.set("")
        self.header_row_var.set("0")
        self.skiprows_var.set("")
        self.delimiter_var.set(",")
        self.encoding_var.set("utf-8")
        self.selection_mode.set("preview")
        self._clear_metadata_selection()
        self._update_info_panel()

    def _set_selection_mode(self) -> None:
        if not hasattr(self, "preview_tree"):
            return
        mode = self.selection_mode.get()
        if mode == "metadata":
            self.preview_tree.configure(selectmode="none")
        else:
            self.preview_tree.configure(selectmode="extended")
            self.preview_tree.selection_remove(self.preview_tree.selection())

    def _clear_metadata_selection(self) -> None:
        self.selected_metadata_cells.clear()
        self.metadata_cells = []
        if hasattr(self, "preview_tree"):
            for item_id in self.preview_tree.get_children():
                self.preview_tree.item(item_id, tags=())
        if hasattr(self, "metadata_listbox") and self.metadata_listbox is not None:
            self.metadata_listbox.delete(0, tk.END)

    def _apply_metadata_tags(self) -> None:
        if not hasattr(self, "preview_tree"):
            return
        row_map: dict[int, bool] = {}
        for row_idx, _col_idx in self.selected_metadata_cells:
            row_map[row_idx] = True
        for item_id in self.preview_tree.get_children():
            row_idx = self.preview_tree.index(item_id)
            if row_map.get(row_idx):
                self.preview_tree.item(item_id, tags=("metadata",))
            else:
                self.preview_tree.item(item_id, tags=())

    def _sync_metadata_cells(self) -> None:
        self.metadata_cells = []
        if self.preview_df is None:
            return
        for row_idx, col_idx in self.selected_metadata_cells:
            if row_idx < 0 or col_idx < 0:
                continue
            if row_idx >= len(self.preview_df.index):
                continue
            if col_idx >= len(self.preview_df.columns):
                continue
            value = self.preview_df.iat[row_idx, col_idx]
            name = "" if value is None else str(value)
            self.metadata_cells.append(
                HeaderCell(
                    name=name,
                    column=col_idx,
                    row=row_idx,
                    alias=None,
                    is_metadata=True,
                    metadata_type="metadata",
                )
            )

    def on_preview_row_select(self, _event=None) -> None:
        """Handle row selection in the data preview to set header row."""
        if self.selection_mode.get() != "preview":
            return
        if not hasattr(self, "preview_tree"):
            return
        selected_items = self.preview_tree.selection()
        if not selected_items:
            return
        item_id = selected_items[0]
        try:
            row_index = self.preview_tree.index(item_id)
            self.header_row_var.set(str(row_index))
            self.load_headers()
        except Exception:
            return

    def _on_cell_click(self, event) -> None:
        if self.selection_mode.get() != "metadata":
            return
        if not hasattr(self, "preview_tree"):
            return
        if not hasattr(self, "metadata_listbox") or self.metadata_listbox is None:
            return
        item_id = self.preview_tree.identify_row(event.y)
        col_id = self.preview_tree.identify_column(event.x)
        if not item_id or not col_id:
            return
        try:
            row_index = self.preview_tree.index(item_id)
            col_index = int(col_id.lstrip("#")) - 1
        except Exception:
            return
        if col_index < 0:
            return

        col_name = f"Col {col_index}"
        cell_value = ""
        if self.preview_df is not None:
            if col_index < len(self.preview_df.columns):
                col_name = str(self.preview_df.columns[col_index])
            if row_index < len(self.preview_df):
                val = self.preview_df.iat[row_index, col_index]
                cell_value = "" if val is None else str(val)

        key = (row_index, col_index)
        if key in self.selected_metadata_cells:
            self.selected_metadata_cells.remove(key)
            for idx in range(self.metadata_listbox.size() - 1, -1, -1):
                entry = self.metadata_listbox.get(idx)
                if entry.startswith(f"Row {row_index}, Col {col_name}"):
                    self.metadata_listbox.delete(idx)
                    break
        else:
            self.selected_metadata_cells.add(key)
            self.header_row_var.set(str(row_index))
            label = f"Row {row_index}, Col {col_name}: '{cell_value[:30]}'"
            self.metadata_listbox.insert(tk.END, label)
            self.metadata_listbox.see(tk.END)

        self._apply_metadata_tags()
        self._sync_metadata_cells()

    def _decrement_header(self) -> None:
        try:
            current = int(self.header_row_var.get())
        except ValueError:
            return
        self.header_row_var.set(str(max(0, current - 1)))
        self.load_headers()

    def _increment_header(self) -> None:
        try:
            current = int(self.header_row_var.get())
        except ValueError:
            return
        self.header_row_var.set(str(current + 1))
        self.load_headers()

    def _parse_combine_on(self) -> List[str]:
        raw = self.combine_on_var.get() or ""
        return [part.strip() for part in raw.split(",") if part.strip()]

    def _parse_dedupe_on(self) -> List[str]:
        raw = self.dedupe_on_var.get() or ""
        return [part.strip() for part in raw.split(",") if part.strip()]

    def _parse_float(self, raw: str | None) -> float | None:
        if not raw:
            return None
        try:
            return float(raw)
        except ValueError:
            return None

    def _on_connection_select(self) -> None:
        if not self.connection_listbox.curselection():
            return
        idx = self.connection_listbox.curselection()[0]
        if idx < len(self.connections):
            self.connection_name_var.set(self.connections[idx].name)

    def preview_connection(self) -> None:
        if not self.connections:
            messagebox.showwarning("No connections", "Add a connection first.")
            return
        conn = self._selected_connection()
        if not conn:
            messagebox.showwarning("No connection selected", "Select a connection to use.")
            return

        table = self.sql_table_var.get().strip()
        query = self.sql_query_var.get().strip()
        if not table and not query:
            messagebox.showwarning("Missing table/query", "Enter a table name or SQL query to preview.")
            return

        try:
            df = fetch_sql_preview(conn, table=table or None, query=query or None, limit=DEFAULT_PREVIEW_ROWS)
        except Exception as exc:
            messagebox.showerror("SQL Preview Failed", str(exc))
            return

        self.source_type = "sql"
        self.file_path = None
        self.columns = list(df.columns)
        self.columns_listbox.delete(0, tk.END)
        for c in self.columns:
            self.columns_listbox.insert(tk.END, c)

        self.preview_df = df
        self.preview_tree.delete(*self.preview_tree.get_children())
        self.preview_tree["columns"] = list(df.columns)
        for col in df.columns:
            self.preview_tree.heading(col, text=col)
            self.preview_tree.column(col, width=120)
        for row in df.itertuples(index=False):
            self.preview_tree.insert("", tk.END, values=list(row))

        self._update_info_panel()

    def _selected_connection(self) -> ConnectionConfig | None:
        target_name = self.connection_name_var.get()
        if not target_name and self.connection_listbox.curselection():
            idx = self.connection_listbox.curselection()[0]
            if idx < len(self.connections):
                target_name = self.connections[idx].name
        return next((c for c in self.connections if c.name == target_name), None)

    def _use_mapped_keys(self) -> None:
        """Prefill combine keys from mapped target fields."""
        if not self.mapping:
            messagebox.showwarning("No mappings", "Map some columns first.")
            return
        targets = list(dict.fromkeys(self.mapping.values()))
        self.combine_keys_var.set(", ".join(targets))

    def test_selected_connection(self) -> None:
        conn = self._selected_connection()
        if not conn:
            messagebox.showwarning("No connection selected", "Select a connection to test.")
            return
        try:
            msg = test_connection(conn)
            messagebox.showinfo("Connection OK", msg)
        except Exception as exc:
            messagebox.showerror("Connection failed", str(exc))

    def open_connection_manager(self) -> None:
        """Simple dialog to add/edit a connection (stored in-memory only)."""
        top = tk.Toplevel(self.root)
        top.title("Connection Manager")
        top.geometry("420x260")

        fields = {
            "Name": tk.StringVar(),
            "Type (e.g., sql, azure, local)": tk.StringVar(value="sql"),
            "Host / URL": tk.StringVar(),
            "Database / Container": tk.StringVar(),
            "User": tk.StringVar(),
            "Password": tk.StringVar(),
            "Driver (e.g., postgresql+psycopg2 or mssql+pyodbc)": tk.StringVar(value="postgresql+psycopg2"),
            "Port": tk.StringVar(),
        }

        for idx, (label, var) in enumerate(fields.items()):
            ttk.Label(top, text=label).grid(row=idx, column=0, padx=8, pady=4, sticky="w")
            show = "*" if "Password" in label else None
            ttk.Entry(top, textvariable=var, show=show).grid(
                row=idx, column=1, padx=8, pady=4, sticky="ew"
            )

        top.columnconfigure(1, weight=1)

        ttk.Label(
            top,
            text="Tip: leave password blank to use env var <NAME>_PASSWORD",
            foreground="gray",
        ).grid(row=len(fields), column=0, columnspan=2, padx=8, pady=(4, 0), sticky="w")
        ttk.Label(
            top,
            text="SQL Server: driver example mssql+pyodbc; requires ODBC Driver 18.",
            foreground="gray",
        ).grid(row=len(fields) + 1, column=0, columnspan=2, padx=8, pady=(2, 8), sticky="w")

        def save_conn() -> None:
            payload = {k: v.get() for k, v in fields.items()}
            if not payload["Name"]:
                messagebox.showerror("Missing name", "Connection name is required.")
                return
            driver_text = payload.get("Driver (e.g., postgresql+psycopg2 or mssql+pyodbc)", "")
            if driver_text and "postgres" in driver_text and payload.get("Host / URL", "").lower() == "localhost":
                messagebox.showinfo(
                    "Driver note",
                    "For Postgres, install driver: pip install sqlalchemy psycopg2-binary",
                )
            if driver_text and "mssql" in driver_text:
                messagebox.showinfo(
                    "Driver note",
                    "For SQL Server, install driver: pip install sqlalchemy pyodbc\n"
                    "Also install Microsoft ODBC Driver 18.",
                )
            cfg = ConnectionConfig(
                name=payload["Name"],
                type=payload["Type (e.g., sql, azure, local)"] or "sql",
                host=payload["Host / URL"] or None,
                database=payload["Database / Container"] or None,
                user=payload["User"] or None,
                password=payload["Password"] or None,
                driver=driver_text or None,
                port=int(payload["Port"]) if payload.get("Port") else None,
            )
            # Replace if same name exists
            self.connections = [c for c in self.connections if c.name != cfg.name]
            self.connections.append(cfg)
            save_connections(self.connections)
            self.connection_name_var.set(cfg.name)
            self._refresh_connection_list()
            top.destroy()

        ttk.Button(top, text="Save", command=save_conn).grid(
            row=len(fields), column=0, padx=8, pady=10
        )
        ttk.Button(top, text="Cancel", command=top.destroy).grid(
            row=len(fields), column=1, padx=8, pady=10, sticky="e"
        )

    def _refresh_connection_list(self) -> None:
        self.connection_listbox.delete(0, tk.END)
        for conn in self.connections:
            name = conn.name
            conn_type = conn.type
            self.connection_listbox.insert(tk.END, f"{name} [{conn_type}]")

    def _load_from_file(self, path: Path):
        tpl = load_template(path)
        self.header_row_var.set(str(tpl.header_row))
        self.combine_sheets_var.set(tpl.combine_sheets)
        self.combine_on_var.set(", ".join(tpl.combine_on) if tpl.combine_on else "")
        self.trim_strings_var.set(tpl.trim_strings)
        self.drop_empty_rows_var.set(tpl.drop_empty_rows)
        self.drop_null_threshold_var.set(
            "" if tpl.drop_null_columns_threshold is None else str(tpl.drop_null_columns_threshold)
        )
        self.dedupe_on_var.set(", ".join(tpl.dedupe_on) if tpl.dedupe_on else "")
        self.strip_thousands_var.set(tpl.strip_thousands)

        if self.source_type == "excel" and self.sheet_names:
            self.sheet_listbox.selection_clear(0, tk.END)
            targets = tpl.sheets if tpl.combine_sheets else (
                [tpl.sheet] if tpl.sheet is not None else []
            )
            for idx, name in enumerate(self.sheet_names):
                if name in targets or str(idx) in [str(t) for t in targets]:
                    self.sheet_listbox.selection_set(idx)

        self.unpivot_var.set(tpl.unpivot)
        self.mapping = tpl.column_mappings

        # Reload columns based on these settings
        if self.source_type == "excel":
            self.load_headers()

        # Reconstruct header cells
        self.header_cells = {h.name: h for h in tpl.headers}
        self._refresh_mapping_view()
        self._update_info_panel()

    def combine_outputs(self) -> None:
        """Run combine-reports logic with current combine settings."""
        mode = self.combine_mode_var.get()
        keys = [k.strip() for k in self.combine_keys_var.get().split(",") if k.strip()]
        how = self.combine_how_var.get()
        strict = self.combine_strict_var.get()
        pattern = self.combine_pattern_var.get() or "*.xlsx"
        input_dir = Path(self.combine_input_dir_var.get() or "data/output")

        if mode == "merge" and not keys:
            messagebox.showwarning("Missing keys", "Provide merge keys (canonical names) before combining.")
            return

        def work():
            if mode == "merge":
                files = sorted(input_dir.glob(pattern))
                if not files:
                    raise ValueError(f"No files found in {input_dir} matching {pattern}")
                missing_files = []
                for f in files:
                    cols = list(read_frame(f).columns)
                    missing = [k for k in keys if k not in cols]
                    if missing:
                        missing_files.append(f"{f.name} (missing: {', '.join(missing)})")
                if missing_files:
                    msg = "Some files are missing merge keys (use canonical names like order_id):\n"
                    msg += "\n".join(missing_files[:5])
                    if len(missing_files) > 5:
                        msg += f"\n...and {len(missing_files)-5} more"
                    raise ValueError(msg)

            df = run_combine(
                input_dir=input_dir,
                pattern=pattern,
                mode=mode,
                keys=keys,
                how=how,
                strict_schema=strict,
            )
            out_path = Path(self.combine_output_var.get() or "data/output/Master_Sales_Report.xlsx")
            out_path.parent.mkdir(parents=True, exist_ok=True)
            if out_path.suffix.lower() == ".parquet":
                df.to_parquet(out_path, index=False)
            else:
                df.to_excel(out_path, index=False)
            return df, out_path

        def on_success(result):
            df, out_path = result
            messagebox.showinfo(
                "Combine complete",
                f"Combined {len(df)} rows using mode={mode}. Saved to {out_path}.",
            )
            self._clear_busy("Combine complete")

        def on_error(exc: Exception):
            messagebox.showerror("Combine failed", str(exc))
            self._clear_busy("Error")

        self._run_worker(work, on_success=on_success, on_error=on_error, message="Combining outputs...")


if __name__ == "__main__":
    root = tk.Tk()
    ExcelTemplateApp(root)
    root.mainloop()
